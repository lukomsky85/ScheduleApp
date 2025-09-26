import sys
import json
import os
import random
import shutil
import zipfile
import itertools
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PyQt5.QtCore import Qt, QTimer, pyqtSignal, QSortFilterProxyModel, QModelIndex
from PyQt5.QtGui import QFont, QIcon, QColor, QPalette, QStandardItemModel, QStandardItem
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QTabWidget, QTableWidget, QTableWidgetItem, QHeaderView, QLabel, QPushButton,
    QComboBox, QSpinBox, QProgressBar, QStatusBar, QFileDialog, QMessageBox,
    QDialog, QLineEdit, QFormLayout, QDialogButtonBox, QGroupBox, QScrollArea,
    QTreeView, QTableView, QAbstractItemView, QMenu, QAction, QDateEdit, QCalendarWidget,
    QFrame, QSizePolicy, QCheckBox, QSpinBox, QDoubleSpinBox, QMenuBar, QToolBar,
    QTextEdit, QListWidget, QListWidgetItem, QInputDialog
)
import calendar
from datetime import datetime as dt_datetime

COLORS = {
    'primary': '#4a6fa5',
    'secondary': '#166088',
    'accent': '#47b8e0',
    'success': '#3cba92',
    'warning': '#f4a261',
    'danger': '#e71d36',
    'light': '#f8f9fa',
    'dark': '#343a40',
    'border': '#dee2e6',
    'hover': '#e9ecef'
}

class BellScheduleEditor(QDialog):
    def __init__(self, parent, current_schedule):
        super().__init__(parent)
        self.setWindowTitle("Редактор расписания звонков")
        self.setModal(True)
        self.resize(500, 400)
        self.current_schedule = current_schedule
        self.result = None
        self.setup_ui()
        self.load_schedule_from_string(current_schedule)
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        title_label = QLabel("Уроки:")
        title_label.setFont(QFont('Segoe UI', 10, QFont.Bold))
        layout.addWidget(title_label)
        self.tree = QTableWidget(0, 3)
        self.tree.setHorizontalHeaderLabels(['№', 'Начало', 'Конец'])
        self.tree.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tree.setSelectionMode(QAbstractItemView.SingleSelection)
        layout.addWidget(self.tree, 1)
        buttons_layout = QHBoxLayout()
        self.add_btn = QPushButton("➕ Добавить")
        self.edit_btn = QPushButton("✏️ Редактировать")
        self.delete_btn = QPushButton("🗑️ Удалить")
        self.save_btn = QPushButton("💾 Сохранить")
        self.edit_btn.setEnabled(False)
        self.delete_btn.setEnabled(False)
        self.add_btn.clicked.connect(self.add_interval)
        self.edit_btn.clicked.connect(self.edit_interval)
        self.delete_btn.clicked.connect(self.delete_interval)
        self.save_btn.clicked.connect(self.save_and_close)
        self.tree.itemSelectionChanged.connect(self.on_select)
        buttons_layout.addWidget(self.add_btn)
        buttons_layout.addWidget(self.edit_btn)
        buttons_layout.addWidget(self.delete_btn)
        buttons_layout.addStretch()
        buttons_layout.addWidget(self.save_btn)
        layout.addLayout(buttons_layout)
    
    def load_schedule_from_string(self, schedule_str):
        if not schedule_str.strip():
            return
        intervals = schedule_str.split(',')
        for i, interval in enumerate(intervals):
            parts = interval.strip().split('-')
            if len(parts) == 2:
                start_time, end_time = parts[0].strip(), parts[1].strip()
                row = self.tree.rowCount()
                self.tree.insertRow(row)
                self.tree.setItem(row, 0, QTableWidgetItem(str(i + 1)))
                self.tree.setItem(row, 1, QTableWidgetItem(start_time))
                self.tree.setItem(row, 2, QTableWidgetItem(end_time))
    
    def add_interval(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавить интервал")
        dialog.setModal(True)
        dialog.resize(300, 150)
        form_layout = QFormLayout(dialog)
        start_var = QLineEdit("8:00")
        end_var = QLineEdit("8:45")
        form_layout.addRow("Начало (ч:мм):", start_var)
        form_layout.addRow("Конец (ч:мм):", end_var)
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(lambda: self._save_add_interval(start_var.text(), end_var.text(), dialog))
        button_box.rejected.connect(dialog.reject)
        form_layout.addRow(button_box)
        dialog.exec_()
    
    def _save_add_interval(self, start_time, end_time, dialog):
        if not start_time or not end_time:
            QMessageBox.warning(self, "Предупреждение", "Введите время начала и конца")
            return
        row = self.tree.rowCount()
        self.tree.insertRow(row)
        self.tree.setItem(row, 0, QTableWidgetItem(str(row + 1)))
        self.tree.setItem(row, 1, QTableWidgetItem(start_time))
        self.tree.setItem(row, 2, QTableWidgetItem(end_time))
        dialog.accept()
    
    def edit_interval(self):
        selected_items = self.tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Предупреждение", "Выберите интервал для редактирования")
            return
        row = selected_items[0].row()
        start_time = self.tree.item(row, 1).text()
        end_time = self.tree.item(row, 2).text()
        dialog = QDialog(self)
        dialog.setWindowTitle("Редактировать интервал")
        dialog.setModal(True)
        dialog.resize(300, 150)
        form_layout = QFormLayout(dialog)
        start_var = QLineEdit(start_time)
        end_var = QLineEdit(end_time)
        form_layout.addRow("Начало (ч:мм):", start_var)
        form_layout.addRow("Конец (ч:мм):", end_var)
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(lambda: self._save_edit_interval(row, start_var.text(), end_var.text(), dialog))
        button_box.rejected.connect(dialog.reject)
        form_layout.addRow(button_box)
        dialog.exec_()
    
    def _save_edit_interval(self, row, start_time, end_time, dialog):
        if not start_time or not end_time:
            QMessageBox.warning(self, "Предупреждение", "Введите время начала и конца")
            return
        self.tree.setItem(row, 1, QTableWidgetItem(start_time))
        self.tree.setItem(row, 2, QTableWidgetItem(end_time))
        dialog.accept()
    
    def delete_interval(self):
        selected_items = self.tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Предупреждение", "Выберите интервал для удаления")
            return
        if QMessageBox.question(self, "Подтверждение", "Вы уверены, что хотите удалить этот интервал?") == QMessageBox.Yes:
            row = selected_items[0].row()
            self.tree.removeRow(row)
            self.renumber_intervals()
    
    def renumber_intervals(self):
        for row in range(self.tree.rowCount()):
            self.tree.setItem(row, 0, QTableWidgetItem(str(row + 1)))
    
    def on_select(self):
        selected = self.tree.selectedItems()
        has_selection = len(selected) > 0
        self.edit_btn.setEnabled(has_selection)
        self.delete_btn.setEnabled(has_selection)
    
    def save_and_close(self):
        intervals = []
        for row in range(self.tree.rowCount()):
            start_time = self.tree.item(row, 1).text()
            end_time = self.tree.item(row, 2).text()
            intervals.append(f"{start_time}-{end_time}")
        self.result = ','.join(intervals)
        self.accept()

class TimeSortProxyModel(QSortFilterProxyModel):
    def lessThan(self, left, right):
        left_data = self.sourceModel().data(left)
        right_data = self.sourceModel().data(right)
        def parse_time(time_str):
            start_time_str = time_str.split('-')[0].strip()
            return dt_datetime.strptime(start_time_str, '%H:%M')
        try:
            left_start = parse_time(left_data)
            right_start = parse_time(right_data)
            return left_start < right_start
        except ValueError:
            return left_data < right_data

# ========================
# УНИВЕРСАЛЬНЫЙ ДИАЛОГ
# ========================
class BaseEntityDialog(QDialog):
    def __init__(self, parent, title, fields, data=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        self.resize(450, 350)
        self.fields = fields
        self.widgets = {}
        self.data = data or {}

        layout = QVBoxLayout(self)
        form_layout = QFormLayout()

        for field in fields:
            key = field['key']
            label = field['label']
            ftype = field['type']
            default = self.data.get(key, field.get('default', ''))

            if ftype == 'text':
                widget = QLineEdit(str(default))
            elif ftype == 'spin':
                widget = QSpinBox()
                widget.setRange(field.get('min', 0), field.get('max', 100))
                widget.setValue(int(default))
            elif ftype == 'combo':
                widget = QComboBox()
                widget.addItems(field['options'])
                if default in field['options']:
                    widget.setCurrentText(default)
            else:
                raise ValueError(f"Unsupported field type: {ftype}")

            self.widgets[key] = widget
            form_layout.addRow(label, widget)

        layout.addLayout(form_layout)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def get_data(self):
        result = {}
        for key, widget in self.widgets.items():
            if isinstance(widget, QLineEdit):
                result[key] = widget.text()
            elif isinstance(widget, QSpinBox):
                result[key] = widget.value()
            elif isinstance(widget, QComboBox):
                result[key] = widget.currentText()
        return result


class ScheduleApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("🎓 Система автоматического составления расписания")
        self.setGeometry(100, 100, 1400, 900)
        self.setMinimumSize(1200, 800)
        self.settings = {
            'days_per_week': 5,
            'lessons_per_day': 6,
            'weeks': 2,
            'start_date': datetime.now().date().isoformat(),
            'school_name': 'Образовательное учреждение',
            'director': 'Директор',
            'academic_year': f"{datetime.now().year}-{datetime.now().year + 1}",
            'auto_backup': True,
            'backup_interval': 30,
            'max_backups': 10,
            'last_academic_year_update': datetime.now().year,
            'bell_schedule': '8:00-8:45,8:55-9:40,9:50-10:35,10:45-11:30,11:40-12:25,12:35-13:20'
        }
        self.groups = []
        self.teachers = []
        self.classrooms = []
        self.subjects = []
        self.schedule = pd.DataFrame()
        self.substitutions = []
        self.holidays = []
        self.backup_timer = None
        self.last_backup_time = None
        self.next_backup_time = None
        self.backup_dir = Path.home() / "AppData" / "Local" / "ScheduleApp" / "backups"
        if not self.backup_dir.exists():
            self.backup_dir.mkdir(parents=True, exist_ok=True)
        self.archive_dir = Path.home() / "AppData" / "Local" / "ScheduleApp" / "schedule_archive"
        if not self.archive_dir.exists():
            self.archive_dir.mkdir(parents=True, exist_ok=True)
        self.create_widgets()
        self.load_data()
        self.start_auto_backup()
        self.check_and_update_experience()

    # ========================
    # УНИВЕРСАЛЬНЫЕ МЕТОДЫ
    # ========================
    def load_table_data(self, table_widget, data_list, columns):
        table_widget.setRowCount(0)
        for item in data_list:
            row = table_widget.rowCount()
            table_widget.insertRow(row)
            for col_idx, col in enumerate(columns):
                value = str(item.get(col['key'], ''))
                table_widget.setItem(row, col_idx, QTableWidgetItem(value))

    def delete_entity(self, entity_list, tree_widget, id_column=0, name_column=1, entity_name="элемент"):
        selected = tree_widget.selectedItems()
        if not selected:
            QMessageBox.information(self, "Информация", f"Выберите {entity_name} для удаления")
            return
        row = selected[0].row()
        entity_id = int(tree_widget.item(row, id_column).text())
        if QMessageBox.question(self, "Подтверждение", f"Удалить {entity_name}?") == QMessageBox.Yes:
            entity_list[:] = [e for e in entity_list if e['id'] != entity_id]
            self.create_backup()

    def open_entity_dialog(self, title, fields, data=None, validate_fn=None):
        dialog = BaseEntityDialog(self, title, fields, data)
        if dialog.exec_() == QDialog.Accepted:
            result = dialog.get_data()
            if validate_fn and not validate_fn(result):
                return None
            return result
        return None

    def create_widgets(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        menubar = self.menuBar()
        file_menu = menubar.addMenu("Файл")
        save_action = QAction("💾 Сохранить", self)
        save_action.triggered.connect(self.save_data)
        load_action = QAction("📂 Загрузить", self)
        load_action.triggered.connect(self.load_data)
        settings_action = QAction("⚙️ Настройки", self)
        settings_action.triggered.connect(self.open_settings)
        backup_action = QAction("🛡️ Бэкап", self)
        backup_action.triggered.connect(self.open_backup_manager)
        about_action = QAction("❓ О программе", self)
        about_action.triggered.connect(self.show_about)
        file_menu.addAction(save_action)
        file_menu.addAction(load_action)
        file_menu.addAction(settings_action)
        file_menu.addAction(backup_action)
        file_menu.addAction(about_action)
        help_menu = menubar.addMenu("Помощь")
        check_update_action = QAction("🔄 Проверить обновления", self)
        check_update_action.triggered.connect(self.check_for_updates)
        help_menu.addAction(check_update_action)
        title_frame = QFrame()
        title_layout = QHBoxLayout(title_frame)
        title_label = QLabel("🎓 Система автоматического составления расписания")
        title_label.setFont(QFont('Segoe UI', 18, QFont.Bold))
        title_label.setStyleSheet(f"color: {COLORS['secondary']};")
        title_layout.addWidget(title_label)
        title_layout.addStretch()
        backup_indicator_frame = QFrame()
        backup_indicator_layout = QVBoxLayout(backup_indicator_frame)
        self.backup_status_label = QLabel("Авто-бэкап: ВКЛ")
        self.backup_status_label.setFont(QFont('Segoe UI', 10, QFont.Bold))
        self.backup_status_label.setStyleSheet(f"color: {COLORS['success']};")
        self.backup_info_label = QLabel("Следующий: --:--")
        self.backup_info_label.setFont(QFont('Segoe UI', 10, QFont.Bold))
        backup_indicator_layout.addWidget(self.backup_status_label)
        backup_indicator_layout.addWidget(self.backup_info_label)
        title_layout.addWidget(backup_indicator_frame)
        main_layout.addWidget(title_frame)

        # Кнопки управления остаются
        # Кнопки управления остаются
        buttons_frame = QFrame()
        buttons_layout = QHBoxLayout(buttons_frame)

        # Создаем кнопки с иконками и текстом
        buttons_config = [
            ("🚀 Сгенерировать расписание", "generate_schedule_thread"),
            ("🔍 Проверить конфликт", "check_conflicts"),
            ("⚡ Оптимизировать", "optimize_schedule"),
            ("📊 Отчеты", "show_reports"),
            ("📤 Экспорт в Excel", "export_to_excel"),
            ("📊 Экспорт по группам", "export_group_schedule_to_excel"),
            ("🌐 Экспорт на сайт", "export_to_website"),
            ("🔄 Журнал замен", "open_substitutions")
        ]

        # Словарь для связи текста кнопок с методами
        button_methods = {
            "🚀 Сгенерировать расписание": self.generate_schedule_thread,
            "🔍 Проверить конфликт": self.check_conflicts,
            "⚡ Оптимизировать": self.optimize_schedule,
            "📊 Отчеты": self.show_reports,
            "📤 Экспорт в Excel": self.export_to_excel,
            "📊 Экспорт по группам": self.export_group_schedule_to_excel,
            "🌐 Экспорт на сайт": self.export_to_website,
            "🔄 Журнал замен": self.open_substitutions
        }

        # Создаем и добавляем кнопки
        for text, method_name in buttons_config:
            btn = QPushButton(text)
            method = button_methods.get(text)
            if method:
                btn.clicked.connect(method)
            else:
                print(f"Внимание: метод {method_name} не найден")
            buttons_layout.addWidget(btn)

        # Добавляем растягивающееся пространство
        buttons_layout.addStretch()

        # Устанавливаем отступы для лучшего внешнего вида
        buttons_layout.setSpacing(10)  # Расстояние между кнопками
        buttons_layout.setContentsMargins(5, 5, 5, 5)  # Отступы вокруг frame

        main_layout.addWidget(buttons_frame)

        # Индикатор прогресса
        self.progress = QProgressBar()
        self.progress.setRange(0, 0)
        self.progress.hide()
        main_layout.addWidget(self.progress)

        # Вкладки
        self.notebook = QTabWidget()
        main_layout.addWidget(self.notebook, 1)
        self.create_groups_tab()
        self.create_teachers_tab()
        self.create_classrooms_tab()
        self.create_subjects_tab()
        self.create_schedule_tab()
        self.create_reports_tab()
        self.create_holidays_tab()
        self.create_archive_tab()

        # Статус-бар
        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)
        self.status_var = "Готов к работе"
        self.statusBar.showMessage(self.status_var)
        self.update_backup_indicator()

    def create_groups_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        btn_frame = QFrame()
        btn_layout = QHBoxLayout(btn_frame)
        add_btn = QPushButton("➕ Добавить группу")
        add_btn.clicked.connect(self.add_group)
        edit_btn = QPushButton("✏️ Редактировать")
        edit_btn.clicked.connect(self.edit_group)
        delete_btn = QPushButton("🗑️ Удалить группу")
        delete_btn.clicked.connect(self.delete_group)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(edit_btn)
        btn_layout.addWidget(delete_btn)
        btn_layout.addStretch()
        layout.addWidget(btn_frame)
        self.groups_tree = QTableWidget(0, 6)
        self.groups_tree.setHorizontalHeaderLabels(['ID', 'Название', 'Тип', 'Студентов', 'Курс', 'Специальность'])
        self.groups_tree.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.groups_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        layout.addWidget(self.groups_tree, 1)
        self.notebook.addTab(tab, "👥 Группы")

    def create_teachers_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        btn_frame = QFrame()
        btn_layout = QHBoxLayout(btn_frame)
        add_btn = QPushButton("➕ Добавить преподавателя")
        add_btn.clicked.connect(self.add_teacher)
        edit_btn = QPushButton("✏️ Редактировать")
        edit_btn.clicked.connect(self.edit_teacher)
        delete_btn = QPushButton("🗑️ Удалить преподавателя")
        delete_btn.clicked.connect(self.delete_teacher)
        update_exp_btn = QPushButton("🔄 Обновить стаж")
        update_exp_btn.clicked.connect(self.update_all_experience)
        recalculate_btn = QPushButton("🔄 Пересчитать часы")
        recalculate_btn.clicked.connect(self.recalculate_teacher_hours)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(edit_btn)
        btn_layout.addWidget(delete_btn)
        btn_layout.addWidget(update_exp_btn)
        btn_layout.addWidget(recalculate_btn)
        btn_layout.addStretch()
        layout.addWidget(btn_frame)
        self.teachers_tree = QTableWidget(0, 11)
        self.teachers_tree.setHorizontalHeaderLabels([
            'ID', 'ФИО',
            'План по предметам', 'Макс. часов',
            'Квалификация', 'Стаж',
            'План всего', 'Факт всего',
            'Факт по предметам', 'Остаток часов', 'Свободных часов'
        ])
        self.teachers_tree.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.teachers_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        layout.addWidget(self.teachers_tree, 1)
        self.notebook.addTab(tab, "👨‍🏫 Преподаватели")

    def create_classrooms_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        btn_frame = QFrame()
        btn_layout = QHBoxLayout(btn_frame)
        add_btn = QPushButton("➕ Добавить аудиторию")
        add_btn.clicked.connect(self.add_classroom)
        edit_btn = QPushButton("✏️ Редактировать")
        edit_btn.clicked.connect(self.edit_classroom)
        delete_btn = QPushButton("🗑️ Удалить аудиторию")
        delete_btn.clicked.connect(self.delete_classroom)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(edit_btn)
        btn_layout.addWidget(delete_btn)
        btn_layout.addStretch()
        layout.addWidget(btn_frame)
        self.classrooms_tree = QTableWidget(0, 6)
        self.classrooms_tree.setHorizontalHeaderLabels(['ID', 'Номер', 'Вместимость', 'Тип', 'Оборудование', 'Расположение'])
        self.classrooms_tree.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.classrooms_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        layout.addWidget(self.classrooms_tree, 1)
        self.notebook.addTab(tab, "🏫 Аудитории")

    def create_subjects_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        btn_frame = QFrame()
        btn_layout = QHBoxLayout(btn_frame)
        add_btn = QPushButton("➕ Добавить предмет")
        add_btn.clicked.connect(self.add_subject)
        edit_btn = QPushButton("✏️ Редактировать")
        edit_btn.clicked.connect(self.edit_subject)
        delete_btn = QPushButton("🗑️ Удалить предмет")
        delete_btn.clicked.connect(self.delete_subject)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(edit_btn)
        btn_layout.addWidget(delete_btn)
        btn_layout.addStretch()
        layout.addWidget(btn_frame)
        self.subjects_tree = QTableWidget(0, 6)
        self.subjects_tree.setHorizontalHeaderLabels(['ID', 'Название', 'Тип группы', 'Часов/неделю', 'Форма контроля', 'Кафедра'])
        self.subjects_tree.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.subjects_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        layout.addWidget(self.subjects_tree, 1)
        self.notebook.addTab(tab, "📚 Предметы")

    def create_schedule_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        filter_frame = QFrame()
        filter_layout = QHBoxLayout(filter_frame)
        filter_layout.addWidget(QLabel("Фильтры:"))
        week_layout = QHBoxLayout()
        week_layout.addWidget(QLabel("Неделя:"))
        self.week_var = QComboBox()
        self.week_var.addItems([f"Неделя {i}" for i in range(1, 13)])
        self.week_var.setCurrentIndex(0)
        self.week_var.currentIndexChanged.connect(self.filter_schedule)
        week_layout.addWidget(self.week_var)
        filter_layout.addLayout(week_layout)
        group_layout = QHBoxLayout()
        group_layout.addWidget(QLabel("Группа:"))
        self.group_filter_var = QComboBox()
        self.group_filter_var.addItem("")
        group_layout.addWidget(self.group_filter_var)
        filter_layout.addLayout(group_layout)
        teacher_layout = QHBoxLayout()
        teacher_layout.addWidget(QLabel("Преподаватель:"))
        self.teacher_filter_var = QComboBox()
        self.teacher_filter_var.addItem("")
        teacher_layout.addWidget(self.teacher_filter_var)
        filter_layout.addLayout(teacher_layout)
        classroom_layout = QHBoxLayout()
        classroom_layout.addWidget(QLabel("Аудитория:"))
        self.classroom_filter_var = QComboBox()
        self.classroom_filter_var.addItem("")
        classroom_layout.addWidget(self.classroom_filter_var)
        filter_layout.addLayout(classroom_layout)
        refresh_btn = QPushButton("🔄 Обновить")
        refresh_btn.clicked.connect(self.filter_schedule)
        filter_layout.addWidget(refresh_btn)
        filter_layout.addStretch()
        layout.addWidget(filter_frame)
        schedule_buttons_frame = QFrame()
        buttons_layout = QHBoxLayout(schedule_buttons_frame)
        buttons = [
            ("➕ Добавить занятие", self.add_lesson),
            ("✏️ Редактировать", self.edit_lesson),
            ("🗑️ Удалить занятие", self.delete_lesson),
            ("🔄 Заменить занятие", self.substitute_lesson),
            ("📅 Календарь", self.show_calendar),
            ("🌐 Экспорт на сайт", self.export_to_website),
            ("⏱️ Найти свободное время", self.find_free_slot),
            ("➡️ Перенести занятие", self.reschedule_lesson)
        ]
        for text, command in buttons:
            btn = QPushButton(text)
            btn.clicked.connect(command)
            buttons_layout.addWidget(btn)
        buttons_layout.addStretch()
        layout.addWidget(schedule_buttons_frame)
        self.schedule_model = QStandardItemModel()
        self.schedule_model.setHorizontalHeaderLabels(['Время', 'Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье'])
        self.schedule_proxy_model = TimeSortProxyModel(self)
        self.schedule_proxy_model.setSourceModel(self.schedule_model)
        self.schedule_view = QTableView()
        self.schedule_view.setModel(self.schedule_proxy_model)
        self.schedule_view.setColumnWidth(0, 100)
        self.schedule_view.verticalHeader().setDefaultSectionSize(100)
        self.schedule_view.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.schedule_view.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.schedule_view, 1)
        self.notebook.addTab(tab, "📅 Расписание")

    def create_reports_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        reports_notebook = QTabWidget()
        layout.addWidget(reports_notebook, 1)
        teacher_report_frame = QWidget()
        teacher_report_layout = QVBoxLayout(teacher_report_frame)
        self.teacher_report_tree = QTableWidget(0, 4)
        self.teacher_report_tree.setHorizontalHeaderLabels(['Преподаватель', 'Часы', 'Группы', 'Предметы'])
        self.teacher_report_tree.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        teacher_report_layout.addWidget(self.teacher_report_tree, 1)
        reports_notebook.addTab(teacher_report_frame, "👨‍🏫 Нагрузка преподавателей")
        group_report_frame = QWidget()
        group_report_layout = QVBoxLayout(group_report_frame)
        self.group_report_tree = QTableWidget(0, 4)
        self.group_report_tree.setHorizontalHeaderLabels(['Группа', 'Часы', 'Предметы', 'Преподаватели'])
        self.group_report_tree.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        group_report_layout.addWidget(self.group_report_tree, 1)
        reports_notebook.addTab(group_report_frame, "👥 Нагрузка групп")
        conflicts_frame = QWidget()
        conflicts_layout = QVBoxLayout(conflicts_frame)
        self.conflicts_text = QTextEdit()
        self.conflicts_text.setReadOnly(True)
        conflicts_layout.addWidget(self.conflicts_text, 1)
        reports_notebook.addTab(conflicts_frame, "⚠️ Конфликты")
        summary_frame = QWidget()
        summary_layout = QVBoxLayout(summary_frame)
        self.summary_text = QTextEdit()
        self.summary_text.setReadOnly(True)
        summary_layout.addWidget(self.summary_text, 1)
        reports_notebook.addTab(summary_frame, "📋 Сводка")
        self.notebook.addTab(tab, "📈 Отчеты")

    def create_holidays_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        btn_frame = QFrame()
        btn_layout = QHBoxLayout(btn_frame)
        add_btn = QPushButton("➕ Добавить праздник")
        add_btn.clicked.connect(self.add_holiday)
        delete_btn = QPushButton("🗑️ Удалить праздник")
        delete_btn.clicked.connect(self.delete_holiday)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(delete_btn)
        btn_layout.addStretch()
        layout.addWidget(btn_frame)
        self.holidays_tree = QTableWidget(0, 3)
        self.holidays_tree.setHorizontalHeaderLabels(['Дата', 'Название', 'Тип'])
        self.holidays_tree.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.holidays_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        layout.addWidget(self.holidays_tree, 1)
        self.notebook.addTab(tab, "🎉 Праздники")

    def create_archive_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Панель кнопок
        btn_frame = QFrame()
        btn_layout = QHBoxLayout(btn_frame)
        
        save_btn = QPushButton("💾 Сохранить текущее расписание")
        save_btn.clicked.connect(self.save_current_schedule)
        
        load_btn = QPushButton("📂 Загрузить расписание")
        load_btn.clicked.connect(self.load_archived_schedule)
        
        delete_btn = QPushButton("🗑️ Удалить")
        delete_btn.clicked.connect(self.delete_archived_schedule)
        
        # НОВАЯ КНОПКА ОБНОВИТЬ
        refresh_btn = QPushButton("🔄 Обновить список")
        refresh_btn.clicked.connect(self.load_archive_list)
        
        export_btn = QPushButton("📤 Экспорт в Excel")
        export_btn.clicked.connect(self.export_archived_schedule)
        
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(load_btn)
        btn_layout.addWidget(delete_btn)
        btn_layout.addWidget(refresh_btn)  # Добавляем кнопку обновить
        btn_layout.addStretch()
        btn_layout.addWidget(export_btn)
        
        layout.addWidget(btn_frame)

        # Таблица архива
        self.archive_tree = QTableWidget(0, 7)
        self.archive_tree.setHorizontalHeaderLabels([
            'Имя файла', 'Дата создания', 'Группы', 
            'Преподаватели', 'Аудитории', 'Предметы', 'Занятий'
        ])
        self.archive_tree.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.archive_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        
        layout.addWidget(self.archive_tree, 1)
        
        # Загружаем данные при создании вкладки
        self.load_archive_list()
        
        self.notebook.addTab(tab, "💾 Архив расписаний")

    # ========================
    # ГРУППЫ — УНИФИЦИРОВАНО
    # ========================
    def add_group(self):
        fields = [
            {'name': 'name', 'label': 'Название:', 'type': 'text', 'key': 'name'},
            {'name': 'type', 'label': 'Тип:', 'type': 'combo', 'key': 'type', 'options': ["основное", "углубленное", "вечернее"]},
            {'name': 'students', 'label': 'Студентов:', 'type': 'spin', 'key': 'students', 'min': 1, 'max': 100, 'default': 25},
            {'name': 'course', 'label': 'Курс:', 'type': 'text', 'key': 'course'},
            {'name': 'specialty', 'label': 'Специальность:', 'type': 'text', 'key': 'specialty'},
        ]
        data = self.open_entity_dialog("Добавить группу", fields, validate_fn=lambda d: bool(d.get('name')))
        if data:
            new_id = max([g['id'] for g in self.groups], default=0) + 1
            data['id'] = new_id
            self.groups.append(data)
            self.load_groups_data()
            self.create_backup()

    def edit_group(self):
        selected = self.groups_tree.selectedItems()
        if not selected:
            return
        row = selected[0].row()
        group_id = int(self.groups_tree.item(row, 0).text())
        group = next((g for g in self.groups if g['id'] == group_id), None)
        if not group:
            return
        fields = [
            {'name': 'name', 'label': 'Название:', 'type': 'text', 'key': 'name'},
            {'name': 'type', 'label': 'Тип:', 'type': 'combo', 'key': 'type', 'options': ["основное", "углубленное", "вечернее"]},
            {'name': 'students', 'label': 'Студентов:', 'type': 'spin', 'key': 'students', 'min': 1, 'max': 100},
            {'name': 'course', 'label': 'Курс:', 'type': 'text', 'key': 'course'},
            {'name': 'specialty', 'label': 'Специальность:', 'type': 'text', 'key': 'specialty'},
        ]
        data = self.open_entity_dialog("Редактировать группу", fields, data=group, validate_fn=lambda d: bool(d.get('name')))
        if data:
            group.update(data)
            self.load_groups_data()
            self.create_backup()

    def delete_group(self):
        self.delete_entity(self.groups, self.groups_tree, entity_name="группу")
        self.load_groups_data()

    def load_groups_data(self):
        columns = [
            {'key': 'id'}, {'key': 'name'}, {'key': 'type'}, {'key': 'students'}, {'key': 'course'}, {'key': 'specialty'}
        ]
        self.load_table_data(self.groups_tree, self.groups, columns)
        self.group_filter_var.clear()
        self.group_filter_var.addItem("")
        for g in self.groups:
            self.group_filter_var.addItem(g['name'])

    # ========================
    # АУДИТОРИИ — УНИФИЦИРОВАНО
    # ========================
    def add_classroom(self):
        fields = [
            {'name': 'name', 'label': 'Номер:', 'type': 'text', 'key': 'name'},
            {'name': 'capacity', 'label': 'Вместимость:', 'type': 'spin', 'key': 'capacity', 'min': 1, 'max': 200, 'default': 30},
            {'name': 'type', 'label': 'Тип:', 'type': 'combo', 'key': 'type', 'options': ["обычная", "компьютерный класс", "спортзал", "лаборатория", "специализированная"]},
            {'name': 'equipment', 'label': 'Оборудование:', 'type': 'text', 'key': 'equipment'},
            {'name': 'location', 'label': 'Расположение:', 'type': 'text', 'key': 'location'},
        ]
        data = self.open_entity_dialog("Добавить аудиторию", fields, validate_fn=lambda d: bool(d.get('name')))
        if data:
            new_id = max([c['id'] for c in self.classrooms], default=0) + 1
            data['id'] = new_id
            self.classrooms.append(data)
            self.load_classrooms_data()
            self.create_backup()

    def edit_classroom(self):
        selected = self.classrooms_tree.selectedItems()
        if not selected:
            return
        row = selected[0].row()
        classroom_id = int(self.classrooms_tree.item(row, 0).text())
        classroom = next((c for c in self.classrooms if c['id'] == classroom_id), None)
        if not classroom:
            return
        fields = [
            {'name': 'name', 'label': 'Номер:', 'type': 'text', 'key': 'name'},
            {'name': 'capacity', 'label': 'Вместимость:', 'type': 'spin', 'key': 'capacity', 'min': 1, 'max': 200},
            {'name': 'type', 'label': 'Тип:', 'type': 'combo', 'key': 'type', 'options': ["обычная", "компьютерный класс", "спортзал", "лаборатория", "специализированная"]},
            {'name': 'equipment', 'label': 'Оборудование:', 'type': 'text', 'key': 'equipment'},
            {'name': 'location', 'label': 'Расположение:', 'type': 'text', 'key': 'location'},
        ]
        data = self.open_entity_dialog("Редактировать аудиторию", fields, data=classroom, validate_fn=lambda d: bool(d.get('name')))
        if data:
            classroom.update(data)
            self.load_classrooms_data()
            self.create_backup()

    def delete_classroom(self):
        self.delete_entity(self.classrooms, self.classrooms_tree, entity_name="аудиторию")
        self.load_classrooms_data()

    def load_classrooms_data(self):
        columns = [
            {'key': 'id'}, {'key': 'name'}, {'key': 'capacity'}, {'key': 'type'}, {'key': 'equipment'}, {'key': 'location'}
        ]
        self.load_table_data(self.classrooms_tree, self.classrooms, columns)
        self.classroom_filter_var.clear()
        self.classroom_filter_var.addItem("")
        for c in self.classrooms:
            self.classroom_filter_var.addItem(c['name'])

    # ========================
    # ПРЕДМЕТЫ — УНИФИЦИРОВАНО
    # ========================
    def add_subject(self):
        fields = [
            {'name': 'name', 'label': 'Название:', 'type': 'text', 'key': 'name'},
            {'name': 'group_type', 'label': 'Тип группы:', 'type': 'combo', 'key': 'group_type', 'options': ["основное", "углубленное", "вечернее"]},
            {'name': 'hours_per_week', 'label': 'Часов/неделю:', 'type': 'spin', 'key': 'hours_per_week', 'min': 1, 'max': 20, 'default': 4},
            {'name': 'assessment', 'label': 'Форма контроля:', 'type': 'combo', 'key': 'assessment', 'options': ["экзамен", "зачет", "зачет с оценкой", "курсовая работа", "дифференцированный зачет"]},
            {'name': 'department', 'label': 'Кафедра:', 'type': 'text', 'key': 'department'},
            {'name': 'description', 'label': 'Описание:', 'type': 'text', 'key': 'description'},
        ]
        data = self.open_entity_dialog("Добавить предмет", fields, validate_fn=lambda d: bool(d.get('name')))
        if data:
            new_id = max([s['id'] for s in self.subjects], default=0) + 1
            data['id'] = new_id
            self.subjects.append(data)
            self.load_subjects_data()
            self.create_backup()

    def edit_subject(self):
        selected = self.subjects_tree.selectedItems()
        if not selected:
            return
        row = selected[0].row()
        subject_id = int(self.subjects_tree.item(row, 0).text())
        subject = next((s for s in self.subjects if s['id'] == subject_id), None)
        if not subject:
            return
        fields = [
            {'name': 'name', 'label': 'Название:', 'type': 'text', 'key': 'name'},
            {'name': 'group_type', 'label': 'Тип группы:', 'type': 'combo', 'key': 'group_type', 'options': ["основное", "углубленное", "вечернее"]},
            {'name': 'hours_per_week', 'label': 'Часов/неделю:', 'type': 'spin', 'key': 'hours_per_week', 'min': 1, 'max': 20},
            {'name': 'assessment', 'label': 'Форма контроля:', 'type': 'combo', 'key': 'assessment', 'options': ["экзамен", "зачет", "зачет с оценкой", "курсовая работа", "дифференцированный зачет"]},
            {'name': 'department', 'label': 'Кафедра:', 'type': 'text', 'key': 'department'},
            {'name': 'description', 'label': 'Описание:', 'type': 'text', 'key': 'description'},
        ]
        data = self.open_entity_dialog("Редактировать предмет", fields, data=subject, validate_fn=lambda d: bool(d.get('name')))
        if data:
            subject.update(data)
            self.load_subjects_data()
            self.create_backup()

    def delete_subject(self):
        self.delete_entity(self.subjects, self.subjects_tree, entity_name="предмет")
        self.load_subjects_data()

    def load_subjects_data(self):
        columns = [
            {'key': 'id'}, {'key': 'name'}, {'key': 'group_type'}, {'key': 'hours_per_week'}, {'key': 'assessment'}, {'key': 'department'}
        ]
        self.load_table_data(self.subjects_tree, self.subjects, columns)

    # ========================
    # ПРЕПОДАВАТЕЛИ — ЗАГРУЗКА ДАННЫХ
    # ========================
    def load_teachers_data(self):
        self.teachers_tree.setRowCount(0)
        for teacher in self.teachers:
            row = self.teachers_tree.rowCount()
            self.teachers_tree.insertRow(row)
            
            # ID и ФИО
            self.teachers_tree.setItem(row, 0, QTableWidgetItem(str(teacher['id'])))
            self.teachers_tree.setItem(row, 1, QTableWidgetItem(teacher['name']))
            
            # План по предметам (многострочный с <br>)
            subject_hours = teacher.get('subject_hours', {})
            plan_subjects_str = ""
            for subj, hours in subject_hours.items():
                plan_subjects_str += f"{subj}: {hours} ч<br>"
            if not plan_subjects_str:
                plan_subjects_str = "—"
            plan_label = QLabel(plan_subjects_str)
            plan_label.setWordWrap(True)
            plan_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
            self.teachers_tree.setCellWidget(row, 2, plan_label)
            
            # Макс. часов
            max_hours = teacher.get('max_hours', 0)
            self.teachers_tree.setItem(row, 3, QTableWidgetItem(str(max_hours)))
            
            # Квалификация и Стаж
            qualification = teacher.get('qualification', '')
            experience = teacher.get('experience', 0)
            self.teachers_tree.setItem(row, 4, QTableWidgetItem(qualification))
            self.teachers_tree.setItem(row, 5, QTableWidgetItem(str(experience)))
            
            # Расчет нагрузки
            load_info = self.calculate_teacher_load(teacher['id'])
            plan_total = load_info['plan']
            fact_total = load_info['fact']
            remaining = load_info['remaining']
            
            self.teachers_tree.setItem(row, 6, QTableWidgetItem(str(plan_total)))
            self.teachers_tree.setItem(row, 7, QTableWidgetItem(str(fact_total)))
            
            # Факт по предметам (многострочный с <br>)
            fact_subjects_str = ""
            fact_by_subject = load_info['fact_by_subject']
            if fact_total > 0 and fact_by_subject:
                for subj, count in fact_by_subject.items():
                    fact_subjects_str += f"{subj}: {count} ч<br>"
            if not fact_subjects_str:
                fact_subjects_str = "—"
            fact_label = QLabel(fact_subjects_str)
            fact_label.setWordWrap(True)
            fact_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
            self.teachers_tree.setCellWidget(row, 8, fact_label)
            
            # Остаток и Свободных часов
            self.teachers_tree.setItem(row, 9, QTableWidgetItem(str(remaining)))
            self.teachers_tree.setItem(row, 10, QTableWidgetItem(str(remaining)))
        
        # Обновление фильтра преподавателей
        self.teacher_filter_var.clear()
        self.teacher_filter_var.addItem("")
        for teacher in self.teachers:
            self.teacher_filter_var.addItem(teacher['name'])
        
        # Подгонка высоты строк под содержимое
        self.teachers_tree.resizeRowsToContents()

    # ========================
    # ПРАЗДНИКИ — ЗАГРУЗКА ДАННЫХ
    # ========================
    def load_holidays_data(self):
        self.holidays_tree.setRowCount(0)
        for holiday in self.holidays:
            row = self.holidays_tree.rowCount()
            self.holidays_tree.insertRow(row)
            self.holidays_tree.setItem(row, 0, QTableWidgetItem(holiday.get('date', '')))
            self.holidays_tree.setItem(row, 1, QTableWidgetItem(holiday.get('name', '')))
            self.holidays_tree.setItem(row, 2, QTableWidgetItem(holiday.get('type', 'Государственный')))

    # ========================
    # АРХИВ — ЗАГРУЗКА ДАННЫХ
    # ========================
    def load_archive_list(self):
        self.archive_tree.setRowCount(0)
        try:
            archive_files = [f for f in os.listdir(self.archive_dir) if f.endswith('.json')]
            archive_files.sort(key=lambda x: os.path.getmtime(os.path.join(self.archive_dir, x)), reverse=True)
            
            for filename in archive_files:
                filepath = os.path.join(self.archive_dir, filename)
                stat = os.stat(filepath)
                creation_time = datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    
                    groups_count = len(data.get('groups', []))
                    teachers_count = len(data.get('teachers', []))
                    classrooms_count = len(data.get('classrooms', []))
                    subjects_count = len(data.get('subjects', []))
                    
                    # Правильно считаем количество занятий
                    schedule = data.get('schedule', [])
                    lessons_count = len([s for s in schedule if s.get('status') in ['подтверждено', 'запланировано']])
                    
                    row = self.archive_tree.rowCount()
                    self.archive_tree.insertRow(row)
                    self.archive_tree.setItem(row, 0, QTableWidgetItem(filename))
                    self.archive_tree.setItem(row, 1, QTableWidgetItem(creation_time))
                    self.archive_tree.setItem(row, 2, QTableWidgetItem(str(groups_count)))
                    self.archive_tree.setItem(row, 3, QTableWidgetItem(str(teachers_count)))
                    self.archive_tree.setItem(row, 4, QTableWidgetItem(str(classrooms_count)))
                    self.archive_tree.setItem(row, 5, QTableWidgetItem(str(subjects_count)))
                    self.archive_tree.setItem(row, 6, QTableWidgetItem(str(lessons_count)))
                    
                except Exception as e:
                    # Упрощенная обработка ошибок
                    row = self.archive_tree.rowCount()
                    self.archive_tree.insertRow(row)
                    self.archive_tree.setItem(row, 0, QTableWidgetItem(filename))
                    self.archive_tree.setItem(row, 1, QTableWidgetItem(creation_time))
                    for i in range(2, 7):
                        self.archive_tree.setItem(row, i, QTableWidgetItem("Ошибка"))
                        
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка загрузки списка архива: {e}")

    # ========================
    # ОСТАЛЬНЫЕ МЕТОДЫ (БЕЗ ИЗМЕНЕНИЙ)
    # ========================

    def fix_teacher_data(self):
        """Исправление проблемных данных преподавателей"""
        for teacher in self.teachers:
            # Исправляем пустую квалификацию
            if not teacher.get('qualification') or teacher.get('qualification') == 'Ошибка':
                teacher['qualification'] = 'не указана'
            
            # Убеждаемся, что все необходимые поля существуют
            if 'subject_hours' not in teacher:
                teacher['subject_hours'] = {}
            if 'max_hours' not in teacher:
                teacher['max_hours'] = 20
            if 'experience' not in teacher:
                teacher['experience'] = 0
        
        self.load_teachers_data()
        QMessageBox.information(self, "Успех", "Данные преподавателей исправлены")

    def recalculate_teacher_hours(self):
        """Пересчитывает нагрузку преподавателей и обновляет таблицу."""
        if not self.teachers:
            QMessageBox.information(self, "Информация", "Нет преподавателей для пересчёта.")
            return

        # Обновляем данные в таблице
        self.load_teachers_data()
        QMessageBox.information(self, "Успех", "Часы преподавателей успешно пересчитаны!")

    def add_teacher(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавить преподавателя")
        dialog.setModal(True)
        dialog.resize(500, 600)
        main_layout = QVBoxLayout(dialog)
        form_layout = QFormLayout()
        name_entry = QLineEdit()
        max_hours_spin = QSpinBox()
        max_hours_spin.setRange(1, 100)
        max_hours_spin.setValue(20)
        plan_hours_spin = QSpinBox()
        plan_hours_spin.setRange(1, 100)
        plan_hours_spin.setValue(0)
        qualification_entry = QLineEdit()
        experience_spin = QSpinBox()
        experience_spin.setRange(0, 50)
        experience_spin.setValue(0)
        contacts_entry = QLineEdit()
        forbidden_days_entry = QLineEdit()
        preferred_days_entry = QLineEdit()
        max_lessons_per_day_spin = QSpinBox()
        max_lessons_per_day_spin.setRange(1, 10)
        max_lessons_per_day_spin.setValue(4)
        form_layout.addRow("ФИО:", name_entry)
        form_layout.addRow("Макс. часов/неделю:", max_hours_spin)
        form_layout.addRow("Часов в плане:", plan_hours_spin)
        form_layout.addRow("Квалификация:", qualification_entry)
        form_layout.addRow("Стаж (лет):", experience_spin)
        form_layout.addRow("Контакты:", contacts_entry)
        form_layout.addRow("Запрещенные дни:", forbidden_days_entry)
        form_layout.addRow("Предпочтительные дни:", preferred_days_entry)
        form_layout.addRow("Макс. пар в день:", max_lessons_per_day_spin)
        main_layout.addLayout(form_layout)
        subjects_label = QLabel("Предметы и часы:")
        main_layout.addWidget(subjects_label)
        subjects_table = QTableWidget(0, 3)
        subjects_table.setHorizontalHeaderLabels(["", "Предмет", "Часов"])
        subjects_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        subjects_table.setColumnWidth(0, 30)
        subjects_table.setColumnWidth(1, 200)
        subjects_table.setColumnWidth(2, 80)
        for subject in self.subjects:
            row = subjects_table.rowCount()
            subjects_table.insertRow(row)
            item = QTableWidgetItem()
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            subjects_table.setItem(row, 0, item)
            subjects_table.setItem(row, 1, QTableWidgetItem(subject['name']))
            hours_spin = QSpinBox()
            hours_spin.setRange(0, 100)
            hours_spin.setValue(0)
            subjects_table.setCellWidget(row, 2, hours_spin)
        main_layout.addWidget(subjects_table)
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(lambda: self._save_teacher(
            name_entry.text(),
            self._get_subject_hours_from_table(subjects_table),
            max_hours_spin.value(),
            plan_hours_spin.value(),
            qualification_entry.text(),
            experience_spin.value(),
            contacts_entry.text(),
            forbidden_days_entry.text(),
            preferred_days_entry.text(),
            max_lessons_per_day_spin.value(),
            dialog
        ))
        button_box.rejected.connect(dialog.reject)
        main_layout.addWidget(button_box)
        dialog.exec_()

    def _get_subject_hours_from_table(self, table_widget):
        subject_hours = {}
        for row in range(table_widget.rowCount()):
            item = table_widget.item(row, 0)
            if item.checkState() == Qt.Checked:
                subject_name = table_widget.item(row, 1).text()
                hours_widget = table_widget.cellWidget(row, 2)
                if isinstance(hours_widget, QSpinBox):
                    hours = hours_widget.value()
                    if hours > 0:
                        subject_hours[subject_name] = hours
        return subject_hours

    def _save_teacher(self, name, subject_hours, max_hours, plan_hours, qualification, experience, contacts, forbidden_days, preferred_days, max_lessons_per_day, dialog):
        if not name:
            QMessageBox.warning(self, "Предупреждение", "Введите ФИО преподавателя")
            return
        new_id = len(self.teachers) + 1
        new_teacher = {
            'id': new_id,
            'name': name,
            'subject_hours': subject_hours,
            'max_hours': max_hours,
            'plan_hours': plan_hours,
            'qualification': qualification,
            'experience': experience,
            'contacts': contacts,
            'forbidden_days': forbidden_days,
            'preferred_days': preferred_days,
            'max_lessons_per_day': max_lessons_per_day
        }
        self.teachers.append(new_teacher)
        self.load_teachers_data()
        self.create_backup()
        dialog.accept()

    def edit_teacher(self):
        selected_items = self.teachers_tree.selectedItems()
        if not selected_items:
            QMessageBox.information(self, "Информация", "Выберите преподавателя для редактирования")
            return
        row = selected_items[0].row()
        teacher_id = int(self.teachers_tree.item(row, 0).text())
        teacher = next((t for t in self.teachers if t['id'] == teacher_id), None)
        if not teacher:
            return
        dialog = QDialog(self)
        dialog.setWindowTitle("Редактировать преподавателя")
        dialog.setModal(True)
        dialog.resize(500, 600)
        main_layout = QVBoxLayout(dialog)
        form_layout = QFormLayout()
        name_entry = QLineEdit(teacher['name'])
        max_hours_spin = QSpinBox()
        max_hours_spin.setRange(1, 100)
        max_hours_spin.setValue(teacher.get('max_hours', 20))
        plan_hours_spin = QSpinBox()
        plan_hours_spin.setRange(1, 100)
        plan_hours_spin.setValue(teacher.get('plan_hours', 0))
        qualification_entry = QLineEdit(teacher.get('qualification', ''))
        experience_spin = QSpinBox()
        experience_spin.setRange(0, 50)
        experience_spin.setValue(teacher.get('experience', 0))
        contacts_entry = QLineEdit(teacher.get('contacts', ''))
        forbidden_days_entry = QLineEdit(teacher.get('forbidden_days', ''))
        preferred_days_entry = QLineEdit(teacher.get('preferred_days', ''))
        max_lessons_per_day_spin = QSpinBox()
        max_lessons_per_day_spin.setRange(1, 10)
        max_lessons_per_day_spin.setValue(teacher.get('max_lessons_per_day', 4))
        form_layout.addRow("ФИО:", name_entry)
        form_layout.addRow("Макс. часов/неделю:", max_hours_spin)
        form_layout.addRow("Часов в плане:", plan_hours_spin)
        form_layout.addRow("Квалификация:", qualification_entry)
        form_layout.addRow("Стаж (лет):", experience_spin)
        form_layout.addRow("Контакты:", contacts_entry)
        form_layout.addRow("Запрещенные дни:", forbidden_days_entry)
        form_layout.addRow("Предпочтительные дни:", preferred_days_entry)
        form_layout.addRow("Макс. пар в день:", max_lessons_per_day_spin)
        main_layout.addLayout(form_layout)
        subjects_label = QLabel("Предметы и часы:")
        main_layout.addWidget(subjects_label)
        subjects_table = QTableWidget(0, 3)
        subjects_table.setHorizontalHeaderLabels(["", "Предмет", "Часов"])
        subjects_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        subjects_table.setColumnWidth(0, 30)
        subjects_table.setColumnWidth(1, 200)
        subjects_table.setColumnWidth(2, 80)
        current_subject_hours = teacher.get('subject_hours', {})
        for subject in self.subjects:
            row = subjects_table.rowCount()
            subjects_table.insertRow(row)
            item = QTableWidgetItem()
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            if subject['name'] in current_subject_hours:
                item.setCheckState(Qt.Checked)
            else:
                item.setCheckState(Qt.Unchecked)
            subjects_table.setItem(row, 0, item)
            subjects_table.setItem(row, 1, QTableWidgetItem(subject['name']))
            hours_spin = QSpinBox()
            hours_spin.setRange(0, 100)
            hours_spin.setValue(current_subject_hours.get(subject['name'], 0))
            subjects_table.setCellWidget(row, 2, hours_spin)
        main_layout.addWidget(subjects_table)
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(lambda: self._update_teacher(
            teacher_id,
            name_entry.text(),
            self._get_subject_hours_from_table(subjects_table),
            max_hours_spin.value(),
            plan_hours_spin.value(),
            qualification_entry.text(),
            experience_spin.value(),
            contacts_entry.text(),
            forbidden_days_entry.text(),
            preferred_days_entry.text(),
            max_lessons_per_day_spin.value(),
            dialog
        ))
        button_box.rejected.connect(dialog.reject)
        main_layout.addWidget(button_box)
        dialog.exec_()

    def _update_teacher(self, teacher_id, name, subject_hours, max_hours, plan_hours, qualification, experience, contacts, forbidden_days, preferred_days, max_lessons_per_day, dialog):
        if not name:
            QMessageBox.warning(self, "Предупреждение", "Введите ФИО преподавателя")
            return
        teacher = next((t for t in self.teachers if t['id'] == teacher_id), None)
        if not teacher:
            return
        teacher['name'] = name
        teacher['subject_hours'] = subject_hours
        teacher['max_hours'] = max_hours
        teacher['plan_hours'] = plan_hours
        teacher['qualification'] = qualification
        teacher['experience'] = experience
        teacher['contacts'] = contacts
        teacher['forbidden_days'] = forbidden_days
        teacher['preferred_days'] = preferred_days
        teacher['max_lessons_per_day'] = max_lessons_per_day
        self.load_teachers_data()
        self.create_backup()
        dialog.accept()

    def delete_teacher(self):
        selected_items = self.teachers_tree.selectedItems()
        if not selected_items:
            QMessageBox.information(self, "Информация", "Выберите преподавателя для удаления")
            return
        if QMessageBox.question(self, "Подтверждение", "Удалить выбранного преподавателя?") == QMessageBox.Yes:
            row = selected_items[0].row()
            teacher_id = int(self.teachers_tree.item(row, 0).text())
            self.teachers = [t for t in self.teachers if t['id'] != teacher_id]
            self.load_teachers_data()
            self.create_backup()

    def update_all_experience(self):
        if QMessageBox.question(self, "Подтверждение", "Вы уверены, что хотите увеличить стаж всех преподавателей на 1 год?") == QMessageBox.Yes:
            updated_count = 0
            for teacher in self.teachers:
                teacher['experience'] = teacher.get('experience', 0) + 1
                updated_count += 1
            self.load_teachers_data()
            self.create_backup()
            QMessageBox.information(self, "Успех", f"Стаж обновлен у {updated_count} преподавателей")

    def generate_schedule_thread(self):
        self.progress.show()
        self.statusBar.showMessage("Генерация расписания...")
        self.generate_schedule()

    def generate_schedule(self):
        if not self.groups or not self.teachers or not self.classrooms or not self.subjects:
            QMessageBox.warning(self, "Предупреждение", "Необходимо добавить группы, преподавателей, аудитории и предметы перед генерацией расписания")
            return
        if not self.settings.get('bell_schedule'):
            QMessageBox.warning(self, "Предупреждение", "Необходимо настроить расписание звонков")
            return
        self.schedule = pd.DataFrame()
        self.create_schedule_structure()
        self.assign_subjects_to_groups()
        self.assign_teachers_and_classrooms()
        self.filter_schedule()
        self.update_reports()
        self.create_backup()
        QMessageBox.information(self, "Успех", "Расписание успешно сгенерировано!")
        # Добавьте эту строку, чтобы скрыть индикатор прогресса
        self.progress.hide()

    def assign_subjects_to_groups(self):
        if not self.subjects or not self.groups:
            return
        for group in self.groups:
            group_subjects = [s for s in self.subjects if s.get('group_type') in [group['type'], 'общий']]
            if not group_subjects:
                continue
            free_slots_count = len(self.schedule[
                (self.schedule['group_id'] == group['id']) &
                (self.schedule['status'] == 'свободно')
            ])
            if free_slots_count == 0:
                continue
            group_subjects.sort(key=lambda x: x.get('hours_per_week', 0), reverse=True)
            for subject in group_subjects:
                hours_per_week = subject.get('hours_per_week', 0)
                if hours_per_week <= 0:
                    continue
                free_slots = self.schedule[
                    (self.schedule['group_id'] == group['id']) &
                    (self.schedule['status'] == 'свободно')
                ].index
                if len(free_slots) < hours_per_week:
                    continue
                selected_slots = random.sample(list(free_slots), hours_per_week)
                for slot in selected_slots:
                    self.schedule.loc[slot, 'subject_id'] = subject['id']
                    self.schedule.loc[slot, 'subject_name'] = subject['name']
                    self.schedule.loc[slot, 'status'] = 'запланировано'

    def create_schedule_structure(self):
        if not self.groups:
            QMessageBox.warning(self, "Ошибка", "Нет групп для распределения расписания")
            return
        days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье'][:self.settings['days_per_week']]
        times = [slot.strip() for slot in self.settings['bell_schedule'].split(',')]
        weeks = list(range(1, self.settings['weeks'] + 1))
        schedule_data = []
        lesson_id = 1
        group_cycle = itertools.cycle(self.groups)
        for week in weeks:
            for day in days:
                for time in times:
                    group = next(group_cycle)
                    schedule_data.append({
                        'id': lesson_id,
                        'week': week,
                        'day': day,
                        'time': time,
                        'group_id': group['id'],
                        'group_name': group['name'],
                        'subject_id': None,
                        'subject_name': '',
                        'teacher_id': None,
                        'teacher_name': '',
                        'classroom_id': None,
                        'classroom_name': '',
                        'status': 'свободно'
                    })
                    lesson_id += 1
        self.schedule = pd.DataFrame(schedule_data)

    def assign_teachers_and_classrooms(self):
        if self.schedule.empty:
            return
        planned_lessons = self.schedule[self.schedule['status'] == 'запланировано'].index
        for idx in planned_lessons:
            subject_name = self.schedule.loc[idx, 'subject_name']
            group_id = self.schedule.loc[idx, 'group_id']
            day = self.schedule.loc[idx, 'day']
            time = self.schedule.loc[idx, 'time']
            week = self.schedule.loc[idx, 'week']
            available_teachers = []
            for teacher in self.teachers:
                if subject_name not in teacher.get('subject_hours', {}):
                    continue
                forbidden_days = teacher.get('forbidden_days', '')
                if day in forbidden_days:
                    continue
                teacher_lessons_today = len(self.schedule[
                    (self.schedule['teacher_id'] == teacher['id']) &
                    (self.schedule['day'] == day) &
                    (self.schedule['status'] == 'подтверждено')
                ])
                if teacher_lessons_today >= teacher.get('max_lessons_per_day', 10):
                    continue
                available_teachers.append(teacher)
            if available_teachers:
                teacher = random.choice(available_teachers)
                self.schedule.loc[idx, 'teacher_id'] = teacher['id']
                self.schedule.loc[idx, 'teacher_name'] = teacher['name']
                self.schedule.loc[idx, 'status'] = 'подтверждено'
                available_classrooms = []
                group_size = self.get_group_size(group_id)
                for classroom in self.classrooms:
                    if classroom.get('capacity', 0) < group_size:
                        continue
                    classroom_lessons = len(self.schedule[
                        (self.schedule['classroom_id'] == classroom['id']) &
                        (self.schedule['day'] == day) &
                        (self.schedule['time'] == time) &
                        (self.schedule['week'] == week) &
                        (self.schedule['status'] == 'подтверждено')
                    ])
                    if classroom_lessons > 0:
                        continue
                    available_classrooms.append(classroom)
                if available_classrooms:
                    classroom = random.choice(available_classrooms)
                    self.schedule.loc[idx, 'classroom_id'] = classroom['id']
                    self.schedule.loc[idx, 'classroom_name'] = classroom['name']
                else:
                    print(f"Предупреждение: Для занятия {subject_name} в {day} {time} (неделя {week}) не найдено подходящей аудитории.")
            else:
                print(f"Предупреждение: Для предмета {subject_name} не найдено подходящего преподавателя.")

    def get_group_size(self, group_id):
        group = next((g for g in self.groups if g['id'] == group_id), None)
        return group.get('students', 0) if group else 0

    def filter_schedule(self):
        current_week_text = self.week_var.currentText()
        try:
            current_week = int(current_week_text.split()[1]) if "Неделя" in current_week_text else 1
        except (ValueError, IndexError):
            current_week = 1
        selected_group = self.group_filter_var.currentText() if self.group_filter_var.currentText() != "" else None
        selected_teacher = self.teacher_filter_var.currentText() if self.teacher_filter_var.currentText() != "" else None
        selected_classroom = self.classroom_filter_var.currentText() if self.classroom_filter_var.currentText() != "" else None
        filtered_schedule = self.schedule.copy()
        filtered_schedule = filtered_schedule[filtered_schedule['week'] == current_week]
        if selected_group:
            filtered_schedule = filtered_schedule[filtered_schedule['group_name'] == selected_group]
        if selected_teacher:
            filtered_schedule = filtered_schedule[filtered_schedule['teacher_name'] == selected_teacher]
        if selected_classroom:
            filtered_schedule = filtered_schedule[filtered_schedule['classroom_name'] == selected_classroom]
        self.schedule_model.clear()
        days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье'][:self.settings['days_per_week']]
        headers = ['Время'] + days
        self.schedule_model.setHorizontalHeaderLabels(headers)
        times = sorted(filtered_schedule['time'].unique(), key=lambda x: x.split('-')[0])
        for time_slot in times:
            time_item = QStandardItem(time_slot)
            self.schedule_model.appendRow([time_item])
        for col, day in enumerate(days, 1):
            for row, time_slot in enumerate(times):
                lesson = filtered_schedule[
                    (filtered_schedule['time'] == time_slot) &
                    (filtered_schedule['day'] == day)
                ]
                if not lesson.empty:
                    lesson_info = lesson.iloc[0]
                    lesson_text = f"{lesson_info['subject_name']}\n{lesson_info['teacher_name']}\n{lesson_info['classroom_name']}"
                    item = QStandardItem(lesson_text)
                    item.setTextAlignment(Qt.AlignCenter)
                    if lesson_info['status'] == 'подтверждено':
                        item.setBackground(QColor('#d4edda'))
                    elif lesson_info['status'] == 'запланировано':
                        item.setBackground(QColor('#fff3cd'))
                    else:
                        item.setBackground(QColor('#999999'))
                    self.schedule_model.setItem(row, col, item)
                else:
                    self.schedule_model.setItem(row, col, QStandardItem())
        self.schedule_proxy_model.setSourceModel(self.schedule_model)
        self.schedule_view.setModel(self.schedule_proxy_model)
        self.schedule_proxy_model.sort(0, Qt.AscendingOrder)
        self.schedule_view.resizeColumnsToContents()
        self.schedule_view.resizeRowsToContents()

    def calculate_teacher_load(self, teacher_id):
        teacher = next((t for t in self.teachers if t['id'] == teacher_id), None)
        if not teacher:
            return {'plan': 0, 'fact': 0, 'remaining': 0, 'fact_by_subject': {}}
        subject_hours = teacher.get('subject_hours', {})
        plan_total = sum(subject_hours.values()) if subject_hours else 0
        fact_total = 0
        fact_by_subject = {}
        if not self.schedule.empty:
            fact_schedule = self.schedule[
                (self.schedule['teacher_id'] == teacher_id) &
                (self.schedule['status'] == 'подтверждено')
            ]
            fact_total = len(fact_schedule)
            if fact_total > 0:
                fact_by_subject = fact_schedule['subject_name'].value_counts().to_dict()
        remaining = max(0, plan_total - fact_total)
        return {
            'plan': plan_total,
            'fact': fact_total,
            'remaining': remaining,
            'fact_by_subject': fact_by_subject
        }

    def add_lesson(self):
        selected_indexes = self.schedule_view.selectionModel().selectedIndexes()
        if not selected_indexes:
            QMessageBox.information(self, "Информация", "Выберите ячейку в таблице для добавления занятия")
            return
        selected_index = selected_indexes[0]
        row = selected_index.row()
        col = selected_index.column()
        if col == 0:
            QMessageBox.warning(self, "Предупреждение", "Выберите ячейку с временем (не первый столбец)")
            return
        time_index = self.schedule_proxy_model.index(row, 0)
        time_slot = self.schedule_proxy_model.data(time_index)
        if not time_slot:
            QMessageBox.warning(self, "Ошибка", "Не удалось определить время занятия")
            return
        day_header = self.schedule_proxy_model.headerData(col, Qt.Horizontal, Qt.DisplayRole)
        if not day_header:
            QMessageBox.warning(self, "Ошибка", "Не удалось определить день недели")
            return
        selected_day = day_header.strip()
        week_text = self.week_var.currentText()
        try:
            selected_week = int(week_text.split()[1]) if "Неделя" in week_text else 1
        except (ValueError, IndexError):
            selected_week = 1
        if not self.groups or not self.subjects or not self.teachers or not self.classrooms:
            QMessageBox.warning(self, "Предупреждение", "Для добавления занятия необходимо, чтобы были созданы хотя бы одна группа, один предмет, один преподаватель и одна аудитория.")
            return
        target_row = self.schedule[
            (self.schedule['week'] == selected_week) &
            (self.schedule['day'] == selected_day) &
            (self.schedule['time'] == time_slot)
        ]
        if target_row.empty:
            QMessageBox.critical(self, "Ошибка", "Не удалось найти подходящий слот в расписании для добавления")
            return
        idx = target_row.index[0]
        if self.schedule.loc[idx, 'status'] != 'свободно':
            if QMessageBox.question(self, "Подтверждение", "В выбранное время уже есть занятие. Заменить его?") != QMessageBox.Yes:
                return
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавить занятие")
        dialog.setModal(True)
        dialog.resize(500, 400)
        form_layout = QFormLayout(dialog)
        group_var = QComboBox()
        group_var.addItems([g['name'] for g in self.groups])
        form_layout.addRow("Группа:", group_var)
        subject_var = QComboBox()
        subject_var.addItems([s['name'] for s in self.subjects])
        form_layout.addRow("Предмет:", subject_var)
        teacher_var = QComboBox()
        teacher_var.addItems([t['name'] for t in self.teachers])
        form_layout.addRow("Преподаватель:", teacher_var)
        classroom_var = QComboBox()
        classroom_var.addItems([c['name'] for c in self.classrooms])
        form_layout.addRow("Аудитория:", classroom_var)
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(lambda: self._save_direct_lesson(idx, group_var.currentText(), subject_var.currentText(), teacher_var.currentText(), classroom_var.currentText(), dialog))
        button_box.rejected.connect(dialog.reject)
        form_layout.addRow(button_box)
        dialog.exec_()

    def _save_direct_lesson(self, idx, group_name, subject_name, teacher_name, classroom_name, dialog):
        selected_group = next((g for g in self.groups if g['name'] == group_name), None)
        selected_subject = next((s for s in self.subjects if s['name'] == subject_name), None)
        selected_teacher = next((t for t in self.teachers if t['name'] == teacher_name), None)
        selected_classroom = next((c for c in self.classrooms if c['name'] == classroom_name), None)
        if not all([selected_group, selected_subject, selected_teacher, selected_classroom]):
            QMessageBox.critical(self, "Ошибка", "Не удалось найти выбранные элементы в базе данных")
            return
        self.schedule.loc[idx, 'group_id'] = selected_group['id']
        self.schedule.loc[idx, 'group_name'] = selected_group['name']
        self.schedule.loc[idx, 'subject_id'] = selected_subject['id']
        self.schedule.loc[idx, 'subject_name'] = selected_subject['name']
        self.schedule.loc[idx, 'teacher_id'] = selected_teacher['id']
        self.schedule.loc[idx, 'teacher_name'] = selected_teacher['name']
        self.schedule.loc[idx, 'classroom_id'] = selected_classroom['id']
        self.schedule.loc[idx, 'classroom_name'] = selected_classroom['name']
        self.schedule.loc[idx, 'status'] = 'подтверждено'
        self.filter_schedule()
        self.update_reports()
        self.create_backup()
        QMessageBox.information(self, "Успех", "Занятие успешно добавлено!")
        dialog.accept()

    def edit_lesson(self):
        selected_indexes = self.schedule_view.selectionModel().selectedIndexes()
        if not selected_indexes:
            QMessageBox.information(self, "Информация", "Выберите занятие в таблице для редактирования")
            return
        selected_index = selected_indexes[0]
        row = selected_index.row()
        col = selected_index.column()
        if col == 0:
            QMessageBox.warning(self, "Предупреждение", "Выберите ячейку с занятием (не время)")
            return
        time_index = self.schedule_proxy_model.index(row, 0)
        time_slot = self.schedule_proxy_model.data(time_index)
        if not time_slot:
            QMessageBox.warning(self, "Ошибка", "Не удалось определить время занятия")
            return
        day_header = self.schedule_proxy_model.headerData(col, Qt.Horizontal, Qt.DisplayRole)
        if not day_header:
            QMessageBox.warning(self, "Ошибка", "Не удалось определить день недели")
            return
        selected_day = day_header.strip()
        week_text = self.week_var.currentText()
        try:
            selected_week = int(week_text.split()[1]) if "Неделя" in week_text else 1
        except (ValueError, IndexError):
            selected_week = 1
        target_lesson = self.schedule[
            (self.schedule['week'] == selected_week) &
            (self.schedule['day'] == selected_day) &
            (self.schedule['time'] == time_slot) &
            (self.schedule['status'] == 'подтверждено')
        ]
        if target_lesson.empty:
            QMessageBox.information(self, "Информация", "Выбранное занятие не найдено в расписании")
            return
        lesson_info = target_lesson.iloc[0]
        lesson_idx = target_lesson.index[0]
        dialog = QDialog(self)
        dialog.setWindowTitle("Редактировать занятие")
        dialog.setModal(True)
        dialog.resize(500, 400)
        form_layout = QFormLayout(dialog)
        week_var = QComboBox()
        week_var.addItems([str(i) for i in range(1, self.settings['weeks'] + 1)])
        week_var.setCurrentText(str(selected_week))
        day_var = QComboBox()
        day_var.addItems(['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье'][:self.settings['days_per_week']])
        day_var.setCurrentText(selected_day)
        time_var = QComboBox()
        bell_schedule_str = self.settings.get('bell_schedule', '8:00-8:45,8:55-9:40,9:50-10:35,10:45-11:30,11:40-12:25,12:35-13:20')
        times = [slot.strip() for slot in bell_schedule_str.split(',')]
        time_var.addItems(times)
        time_var.setCurrentText(time_slot)
        group_var = QComboBox()
        group_var.addItems([g['name'] for g in self.groups])
        group_var.setCurrentText(lesson_info['group_name'])
        subject_var = QComboBox()
        subject_var.addItems([s['name'] for s in self.subjects])
        subject_var.setCurrentText(lesson_info['subject_name'])
        teacher_var = QComboBox()
        teacher_var.addItems([t['name'] for t in self.teachers])
        teacher_var.setCurrentText(lesson_info['teacher_name'])
        classroom_var = QComboBox()
        classroom_var.addItems([c['name'] for c in self.classrooms])
        classroom_var.setCurrentText(lesson_info['classroom_name'])
        form_layout.addRow("Неделя:", week_var)
        form_layout.addRow("День недели:", day_var)
        form_layout.addRow("Время:", time_var)
        form_layout.addRow("Группа:", group_var)
        form_layout.addRow("Предмет:", subject_var)
        form_layout.addRow("Преподаватель:", teacher_var)
        form_layout.addRow("Аудитория:", classroom_var)
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        form_layout.addRow(button_box)
        if dialog.exec_() == QDialog.Accepted:
            new_week = int(week_var.currentText())
            new_day = day_var.currentText()
            new_time = time_var.currentText()
            new_group_name = group_var.currentText()
            new_subject_name = subject_var.currentText()
            new_teacher_name = teacher_var.currentText()
            new_classroom_name = classroom_var.currentText()
            new_group = next((g for g in self.groups if g['name'] == new_group_name), None)
            new_subject = next((s for s in self.subjects if s['name'] == new_subject_name), None)
            new_teacher = next((t for t in self.teachers if t['name'] == new_teacher_name), None)
            new_classroom = next((c for c in self.classrooms if c['name'] == new_classroom_name), None)
            if not all([new_group, new_subject, new_teacher, new_classroom]):
                QMessageBox.critical(self, "Ошибка", "Не удалось найти выбранные элементы в базе данных")
                return
            existing_lesson = self.schedule[
                (self.schedule['week'] == new_week) &
                (self.schedule['day'] == new_day) &
                (self.schedule['time'] == new_time) &
                (self.schedule['group_id'] == new_group['id']) &
                (self.schedule['status'] != 'свободно') &
                (self.schedule.index != lesson_idx)
            ]
            if not existing_lesson.empty:
                if QMessageBox.question(self, "Подтверждение", "В выбранное время у этой группы уже есть занятие. Заменить его?") != QMessageBox.Yes:
                    return
            target_row = self.schedule[
                (self.schedule['week'] == new_week) &
                (self.schedule['day'] == new_day) &
                (self.schedule['time'] == new_time) &
                (self.schedule['group_id'] == new_group['id'])
            ]
            if not target_row.empty:
                update_idx = target_row.index[0]
            else:
                QMessageBox.critical(self, "Ошибка", "Не удалось найти подходящий слот в расписании для обновления")
                return
            self.schedule.loc[update_idx, 'week'] = new_week
            self.schedule.loc[update_idx, 'day'] = new_day
            self.schedule.loc[update_idx, 'time'] = new_time
            self.schedule.loc[update_idx, 'group_id'] = new_group['id']
            self.schedule.loc[update_idx, 'group_name'] = new_group['name']
            self.schedule.loc[update_idx, 'subject_id'] = new_subject['id']
            self.schedule.loc[update_idx, 'subject_name'] = new_subject['name']
            self.schedule.loc[update_idx, 'teacher_id'] = new_teacher['id']
            self.schedule.loc[update_idx, 'teacher_name'] = new_teacher['name']
            self.schedule.loc[update_idx, 'classroom_id'] = new_classroom['id']
            self.schedule.loc[update_idx, 'classroom_name'] = new_classroom['name']
            self.schedule.loc[update_idx, 'status'] = 'подтверждено'
            if update_idx != lesson_idx:
                self.schedule.loc[lesson_idx, ['subject_id', 'subject_name', 'teacher_id', 'teacher_name', 'classroom_id', 'classroom_name']] = [None, '', None, '', None, '']
                self.schedule.loc[lesson_idx, 'status'] = 'свободно'
            self.filter_schedule()
            self.update_reports()
            self.create_backup()
            QMessageBox.information(self, "Успех", "Занятие успешно отредактировано!")

    def delete_lesson(self):
        selected_indexes = self.schedule_view.selectionModel().selectedIndexes()
        if not selected_indexes:
            QMessageBox.information(self, "Информация", "Выберите занятие в таблице для удаления")
            return
        selected_index = selected_indexes[0]
        row = selected_index.row()
        col = selected_index.column()
        if col == 0:
            QMessageBox.warning(self, "Предупреждение", "Выберите ячейку с занятием (не время)")
            return
        cell_data = self.schedule_proxy_model.data(selected_index)
        if not cell_data or not cell_data.strip():
            QMessageBox.warning(self, "Предупреждение", "Выбранная ячейка пуста. Нет занятия для удаления.")
            return
        time_index = self.schedule_proxy_model.index(row, 0)
        time_slot = self.schedule_proxy_model.data(time_index)
        if not time_slot:
            QMessageBox.warning(self, "Ошибка", "Не удалось определить время занятия")
            return
        day_header = self.schedule_proxy_model.headerData(col, Qt.Horizontal, Qt.DisplayRole)
        if not day_header:
            QMessageBox.warning(self, "Ошибка", "Не удалось определить день недели")
            return
        selected_day = day_header.strip()
        week_text = self.week_var.currentText()
        try:
            selected_week = int(week_text.split()[1]) if "Неделя" in week_text else 1
        except (ValueError, IndexError):
            selected_week = 1
        target_lesson = self.schedule[
            (self.schedule['week'] == selected_week) &
            (self.schedule['day'] == selected_day) &
            (self.schedule['time'] == time_slot) &
            (self.schedule['status'] == 'подтверждено')
        ]
        if target_lesson.empty:
            QMessageBox.information(self, "Информация", "Выбранное занятие не найдено в расписании")
            return
        lesson_info = target_lesson.iloc[0]
        confirm_text = (
            f"Вы уверены, что хотите удалить это занятие?\n"
            f"Предмет: {lesson_info['subject_name']}\n"
            f"Группа: {lesson_info['group_name']}\n"
            f"Преподаватель: {lesson_info['teacher_name']}\n"
            f"Аудитория: {lesson_info['classroom_name']}\n"
            f"День: {selected_day}\n"
            f"Время: {time_slot}\n"
            f"Неделя: {selected_week}"
        )
        if QMessageBox.question(self, "Подтверждение удаления", confirm_text) != QMessageBox.Yes:
            return
        idx = target_lesson.index[0]
        self.schedule.loc[idx, ['subject_id', 'subject_name', 'teacher_id', 'teacher_name', 'classroom_id', 'classroom_name']] = [None, '', None, '', None, '']
        self.schedule.loc[idx, 'status'] = 'свободно'
        self.filter_schedule()
        self.update_reports()
        self.create_backup()
        QMessageBox.information(self, "Успех", "Занятие успешно удалено!")

    def substitute_lesson(self):
        selected_indexes = self.schedule_view.selectionModel().selectedIndexes()
        if not selected_indexes:
            QMessageBox.information(self, "Информация", "Выберите занятие в таблице для замены")
            return
        selected_index = selected_indexes[0]
        row = selected_index.row()
        col = selected_index.column()
        if col == 0:
            QMessageBox.warning(self, "Предупреждение", "Выберите ячейку с занятием (не время)")
            return
        cell_data = self.schedule_proxy_model.data(selected_index)
        if not cell_data or not cell_data.strip():
            QMessageBox.warning(self, "Предупреждение", "Выбранная ячейка пуста. Нет занятия для замены.")
            return
        time_index = self.schedule_proxy_model.index(row, 0)
        time_slot = self.schedule_proxy_model.data(time_index)
        if not time_slot:
            QMessageBox.warning(self, "Ошибка", "Не удалось определить время занятия")
            return
        day_header = self.schedule_proxy_model.headerData(col, Qt.Horizontal, Qt.DisplayRole)
        if not day_header:
            QMessageBox.warning(self, "Ошибка", "Не удалось определить день недели")
            return
        selected_day = day_header.strip()
        week_text = self.week_var.currentText()
        try:
            selected_week = int(week_text.split()[1]) if "Неделя" in week_text else 1
        except (ValueError, IndexError):
            selected_week = 1
        target_lesson = self.schedule[
            (self.schedule['week'] == selected_week) &
            (self.schedule['day'] == selected_day) &
            (self.schedule['time'] == time_slot) &
            (self.schedule['status'] == 'подтверждено')
        ]
        if target_lesson.empty:
            QMessageBox.information(self, "Информация", "Выбранное занятие не найдено в расписании")
            return
        lesson_info = target_lesson.iloc[0]
        lesson_idx = target_lesson.index[0]
        dialog = QDialog(self)
        dialog.setWindowTitle("Замена занятия")
        dialog.setModal(True)
        dialog.resize(550, 450)
        layout = QVBoxLayout(dialog)
        info_label = QLabel(f"""
        <b>Замена для занятия:</b>
        <br>Текущий предмет: {lesson_info['subject_name']}
        <br>Группа: {lesson_info['group_name']}
        <br>День: {selected_day}
        <br>Время: {time_slot}
        <br>Текущий преподаватель: {lesson_info['teacher_name']}
        """)
        info_label.setWordWrap(True)
        info_label.setAlignment(Qt.AlignLeft)
        layout.addWidget(info_label)
        subject_label = QLabel("Новый предмет:")
        layout.addWidget(subject_label)
        subject_combo = QComboBox()
        current_group = next((g for g in self.groups if g['name'] == lesson_info['group_name']), None)
        if current_group:
            available_subjects = [s for s in self.subjects if s.get('group_type') in [current_group['type'], 'общий']]
        else:
            available_subjects = self.subjects
        subject_names = [s['name'] for s in available_subjects]
        subject_combo.addItems(subject_names)
        if lesson_info['subject_name'] in subject_names:
            subject_combo.setCurrentText(lesson_info['subject_name'])
        layout.addWidget(subject_combo)
        teacher_label = QLabel("Новый преподаватель:")
        layout.addWidget(teacher_label)
        teacher_combo = QComboBox()
        def update_teacher_combo():
            selected_subject_name = subject_combo.currentText()
            selected_subject = next((s for s in self.subjects if s['name'] == selected_subject_name), None)
            if selected_subject:
                new_available_teachers = [
                    t for t in self.teachers
                    if selected_subject_name in t.get('subject_hours', {})
                ]
                teacher_combo.clear()
                teacher_names = [t['name'] for t in new_available_teachers]
                teacher_combo.addItems(teacher_names)
                if lesson_info['teacher_name'] in teacher_names:
                    teacher_combo.setCurrentText(lesson_info['teacher_name'])
                elif teacher_names:
                    teacher_combo.setCurrentIndex(0)
            else:
                teacher_combo.clear()
        update_teacher_combo()
        subject_combo.currentTextChanged.connect(update_teacher_combo)
        layout.addWidget(teacher_combo)
        reason_label = QLabel("Причина замены:")
        layout.addWidget(reason_label)
        reason_entry = QTextEdit()
        reason_entry.setPlaceholderText("Укажите причину замены (болезнь, замена предмета и т.д.)")
        reason_entry.setMaximumHeight(80)
        layout.addWidget(reason_entry)
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)
        if dialog.exec_() == QDialog.Accepted:
            new_subject_name = subject_combo.currentText()
            new_teacher_name = teacher_combo.currentText()
            reason_text = reason_entry.toPlainText().strip()
            if not reason_text:
                QMessageBox.warning(self, "Предупреждение", "Пожалуйста, укажите причину замены.")
                return
            new_subject = next((s for s in self.subjects if s['name'] == new_subject_name), None)
            if not new_subject:
                QMessageBox.warning(self, "Ошибка", f"Не удалось найти предмет '{new_subject_name}'")
                return
            new_teacher = next((t for t in self.teachers if t['name'] == new_teacher_name and new_subject_name in t.get('subject_hours', {})), None)
            if not new_teacher:
                QMessageBox.warning(self, "Ошибка", f"Не удалось найти преподавателя '{new_teacher_name}', который ведет предмет '{new_subject_name}'")
                return
            self.schedule.loc[lesson_idx, 'subject_id'] = new_subject['id']
            self.schedule.loc[lesson_idx, 'subject_name'] = new_subject['name']
            self.schedule.loc[lesson_idx, 'teacher_id'] = new_teacher['id']
            self.schedule.loc[lesson_idx, 'teacher_name'] = new_teacher['name']
            self.substitutions.append({
                'date': datetime.now().isoformat(),
                'week': selected_week,
                'day': selected_day,
                'time': time_slot,
                'group': lesson_info['group_name'],
                'old_subject': lesson_info['subject_name'],
                'new_subject': new_subject_name,
                'old_teacher': lesson_info['teacher_name'],
                'new_teacher': new_teacher_name,
                'reason': reason_text
            })
            self.filter_schedule()
            self.update_reports()
            self.create_backup()
            QMessageBox.information(self, "Успех", f"Занятие успешно заменено!\nНовый преподаватель: {new_teacher_name}")

    def show_calendar(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Календарь расписания")
        dialog.setModal(True)
        dialog.resize(600, 500)
        layout = QVBoxLayout(dialog)
        calendar = QCalendarWidget()
        layout.addWidget(calendar)
        button_box = QDialogButtonBox(QDialogButtonBox.Ok)
        button_box.accepted.connect(dialog.accept)
        layout.addWidget(button_box)
        dialog.exec_()

    def check_conflicts(self):
        if self.schedule.empty:
            QMessageBox.information(self, "Информация", "Сначала сгенерируйте расписание")
            return
        teacher_conflicts = pd.DataFrame()
        classroom_conflicts = pd.DataFrame()
        group_conflicts = pd.DataFrame()
        if 'teacher_id' in self.schedule.columns:
            teacher_conflicts = self.schedule[
                (self.schedule['status'] == 'подтверждено') &
                (self.schedule.duplicated(['teacher_id', 'day', 'time', 'week'], keep=False))
            ]
        if 'classroom_id' in self.schedule.columns:
            classroom_conflicts = self.schedule[
                (self.schedule['status'] == 'подтверждено') &
                (self.schedule.duplicated(['classroom_id', 'day', 'time', 'week'], keep=False))
            ]
        if 'group_id' in self.schedule.columns:
            group_conflicts = self.schedule[
                (self.schedule['status'] == 'подтверждено') &
                (self.schedule.duplicated(['group_id', 'day', 'time', 'week'], keep=False))
            ]
        conflict_text = f"Конфликты преподавателей: {len(teacher_conflicts)}\n"
        conflict_text += f"Конфликты аудиторий: {len(classroom_conflicts)}\n"
        conflict_text += f"Конфликты групп: {len(group_conflicts)}\n"
        if not teacher_conflicts.empty:
            conflict_text += "Конфликты преподавателей:\n"
            for _, conflict in teacher_conflicts.head(10).iterrows():
                conflict_text += f"  {conflict['teacher_name']} - {conflict['day']} {conflict['time']} (неделя {conflict['week']})\n"
        if not classroom_conflicts.empty:
            conflict_text += "\nКонфликты аудиторий:\n"
            for _, conflict in classroom_conflicts.head(10).iterrows():
                conflict_text += f"  {conflict['classroom_name']} - {conflict['day']} {conflict['time']} (неделя {conflict['week']})\n"
        if not group_conflicts.empty:
            conflict_text += "\nКонфликты групп:\n"
            for _, conflict in group_conflicts.head(10).iterrows():
                conflict_text += f"  {conflict['group_name']} - {conflict['day']} {conflict['time']} (неделя {conflict['week']})\n"
        self.conflicts_text.setText(conflict_text)
        QMessageBox.information(self, "Проверка конфликтов", f"Найдено конфликтов:\nПреподавателей: {len(teacher_conflicts)}\nАудиторий: {len(classroom_conflicts)}\nГрупп: {len(group_conflicts)}")

    def optimize_schedule(self):
        if self.schedule.empty:
            QMessageBox.information(self, "Информация", "Сначала сгенерируйте расписание")
            return
        self.progress.show()
        self.statusBar.showMessage("Оптимизация расписания...")
        conflicts = pd.DataFrame()
        if 'teacher_id' in self.schedule.columns:
            conflicts = self.schedule[
                (self.schedule['status'] == 'подтверждено') &
                (self.schedule.duplicated(['teacher_id', 'day', 'time', 'week'], keep=False))
            ]
        optimized_count = 0
        for idx in conflicts.index[:10]:
            group_id = self.schedule.loc[idx, 'group_id']
            week = self.schedule.loc[idx, 'week']
            free_slots = self.schedule[
                (self.schedule['group_id'] == group_id) &
                (self.schedule['week'] == week) &
                (self.schedule['status'] == 'свободно')
            ].index
            if len(free_slots) > 0:
                new_slot = random.choice(free_slots)
                self.schedule.loc[new_slot, 'subject_id'] = self.schedule.loc[idx, 'subject_id']
                self.schedule.loc[new_slot, 'subject_name'] = self.schedule.loc[idx, 'subject_name']
                self.schedule.loc[new_slot, 'teacher_id'] = self.schedule.loc[idx, 'teacher_id']
                self.schedule.loc[new_slot, 'teacher_name'] = self.schedule.loc[idx, 'teacher_name']
                self.schedule.loc[new_slot, 'classroom_id'] = self.schedule.loc[idx, 'classroom_id']
                self.schedule.loc[new_slot, 'classroom_name'] = self.schedule.loc[idx, 'classroom_name']
                self.schedule.loc[new_slot, 'status'] = 'подтверждено'
                self.schedule.loc[idx, ['subject_id', 'subject_name', 'teacher_id', 'teacher_name', 'classroom_id', 'classroom_name']] = [None, '', None, '', None, '']
                self.schedule.loc[idx, 'status'] = 'свободно'
                optimized_count += 1
        self.progress.hide()
        self.statusBar.showMessage("Оптимизация завершена")
        QMessageBox.information(self, "Оптимизация", f"Оптимизировано {optimized_count} занятий")
        self.filter_schedule()
        self.update_reports()
        self.create_backup()

    def update_reports(self):
        if self.schedule.empty:
            return
        self.teacher_report_tree.setRowCount(0)
        self.group_report_tree.setRowCount(0)
        self.summary_text.clear()
        if 'teacher_name' in self.schedule.columns and not self.schedule[self.schedule['status'] == 'подтверждено'].empty:
            teacher_load = self.schedule[self.schedule['status'] == 'подтверждено'].groupby('teacher_name').size()
            for teacher, hours in teacher_load.items():
                teacher_data = self.schedule[
                    (self.schedule['teacher_name'] == teacher) &
                    (self.schedule['status'] == 'подтверждено')
                ]
                groups = ', '.join(teacher_data['group_name'].unique())
                subjects = ', '.join(teacher_data['subject_name'].unique())
                row = self.teacher_report_tree.rowCount()
                self.teacher_report_tree.insertRow(row)
                self.teacher_report_tree.setItem(row, 0, QTableWidgetItem(teacher))
                self.teacher_report_tree.setItem(row, 1, QTableWidgetItem(str(hours)))
                self.teacher_report_tree.setItem(row, 2, QTableWidgetItem(groups))
                self.teacher_report_tree.setItem(row, 3, QTableWidgetItem(subjects))
        if 'group_name' in self.schedule.columns and not self.schedule[self.schedule['status'] == 'подтверждено'].empty:
            group_load = self.schedule[self.schedule['status'] == 'подтверждено'].groupby('group_name').size()
            for group, hours in group_load.items():
                group_data = self.schedule[
                    (self.schedule['group_name'] == group) &
                    (self.schedule['status'] == 'подтверждено')
                ]
                subjects = ', '.join(group_data['subject_name'].unique())
                teachers = ', '.join(group_data['teacher_name'].unique())
                row = self.group_report_tree.rowCount()
                self.group_report_tree.insertRow(row)
                self.group_report_tree.setItem(row, 0, QTableWidgetItem(group))
                self.group_report_tree.setItem(row, 1, QTableWidgetItem(str(hours)))
                self.group_report_tree.setItem(row, 2, QTableWidgetItem(subjects))
                self.group_report_tree.setItem(row, 3, QTableWidgetItem(teachers))
        summary_text = f"📊 Сводный отчет по расписанию\n"
        summary_text += f"Учреждение: {self.settings.get('school_name', 'Не указано')}\n"
        summary_text += f"Директор: {self.settings.get('director', 'Не указано')}\n"
        summary_text += f"Учебный год: {self.settings.get('academic_year', 'Не указано')}\n"
        summary_text += f"Дата начала: {self.settings.get('start_date', 'Не указано')}\n"
        summary_text += f"📅 Расписание на {self.settings.get('weeks', 0)} недель\n"
        summary_text += f"📅 Дней в неделю: {self.settings.get('days_per_week', 0)}\n"
        summary_text += f"📅 Занятий в день: {self.settings.get('lessons_per_day', 0)}\n"
        summary_text += f"👥 Групп: {len(self.groups)}\n"
        summary_text += f"👨‍🏫 Преподавателей: {len(self.teachers)}\n"
        summary_text += f"🏫 Аудиторий: {len(self.classrooms)}\n"
        summary_text += f"📚 Предметов: {len(self.subjects)}\n"
        summary_text += f"🎉 Праздничных дней: {len(self.holidays)}\n"
        if not self.schedule.empty:
            confirmed_lessons = len(self.schedule[self.schedule['status'] == 'подтверждено'])
            planned_lessons = len(self.schedule[self.schedule['status'] == 'запланировано'])
            free_slots = len(self.schedule[self.schedule['status'] == 'свободно'])
            summary_text += f"✅ Подтвержденных занятий: {confirmed_lessons}\n"
            summary_text += f"📝 Запланированных занятий: {planned_lessons}\n"
            summary_text += f"🕒 Свободных слотов: {free_slots}\n"
        self.summary_text.setText(summary_text)

    def show_reports(self):
        self.notebook.setCurrentIndex(5)
        self.update_reports()

    def export_to_excel(self):
        if self.schedule.empty:
            QMessageBox.information(self, "Информация", "Сначала сгенерируйте расписание")
            return
        filename, _ = QFileDialog.getSaveFileName(self, "Сохранить расписание как", "", "Excel files (*.xlsx);;All files (*.*)")
        if not filename:
            return
        try:
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            # Создаём книгу вручную
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            # === 1. Лист "Расписание" ===
            ws_schedule = wb.create_sheet('Расписание')
            confirmed_schedule = self.schedule[self.schedule['status'] == 'подтверждено']
            schedule_df = confirmed_schedule if not confirmed_schedule.empty else self.schedule

            # Заголовки и стили (остается без изменений)
            schedule_headers = ['Неделя', 'День', 'Время', 'Группа', 'Предмет', 'Преподаватель', 'Аудитория']
            ws_schedule.append(schedule_headers)

            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for col_num in range(1, len(schedule_headers) + 1):
                cell = ws_schedule.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for _, row in schedule_df.iterrows():
                ws_schedule.append([
                    row['week'], row['day'], row['time'], row['group_name'],
                    row['subject_name'], row['teacher_name'], row['classroom_name']
                ])

            for row in ws_schedule.iter_rows(min_row=2, max_row=ws_schedule.max_row, min_col=1, max_col=len(schedule_headers)):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                    cell.border = thin_border

            for col in ws_schedule.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws_schedule.column_dimensions[column].width = adjusted_width

            # === 2. Остальные листы ===
            def export_simple_sheet(data, sheet_name, columns=None):
                if not data:
                    df = pd.DataFrame(columns=columns) if columns else pd.DataFrame()
                else:
                    df = pd.DataFrame(data)
                ws = wb.create_sheet(sheet_name)
                if not df.empty:
                    ws.append(df.columns.tolist())
                    for _, row in df.iterrows():
                        ws.append(row.tolist())

            export_simple_sheet(self.groups, 'Группы')
            export_simple_sheet(self.teachers, 'Преподаватели')
            export_simple_sheet(self.classrooms, 'Аудитории')
            export_simple_sheet(self.subjects, 'Предметы')
            export_simple_sheet(self.holidays, 'Праздники')

            # === 3. Отчёты: Нагрузка преподавателей и групп ===
            if not self.schedule.empty and not self.schedule[self.schedule['status'] == 'подтверждено'].empty:
                # --- Нагрузка преподавателей ---
                teacher_report_data = []
                for teacher in self.teachers:
                    teacher_name = teacher['name']
                    teacher_schedule = self.schedule[
                        (self.schedule['teacher_name'] == teacher_name) &
                        (self.schedule['status'] == 'подтверждено')
                    ]
                    
                    if not teacher_schedule.empty:
                        groups = ', '.join(sorted(teacher_schedule['group_name'].unique()))
                        hours = len(teacher_schedule)

                        # Формируем строку с предметами и часами (используем переносы строк)
                        subject_hours = teacher_schedule['subject_name'].value_counts().to_dict()
                        fact_subjects_str = ""
                        for subj, count in subject_hours.items():
                            fact_subjects_str += f"{subj}: {count} ч\n"  # Используем \n для Excel
                        
                        if not fact_subjects_str:
                            fact_subjects_str = "—"
                        
                        teacher_report_data.append([teacher_name, groups, fact_subjects_str, hours])

                if teacher_report_data:
                    ws_teachers = wb.create_sheet('Нагрузка преподавателей')
                    ws_teachers.append(['Преподаватель', 'Группы', 'Предметы', 'Часы'])
                    
                    for col_num in range(1, 5):
                        cell = ws_teachers.cell(row=1, column=col_num)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    for row_data in teacher_report_data:
                        ws_teachers.append(row_data)
                    
                    for row in ws_teachers.iter_rows(min_row=2, max_row=ws_teachers.max_row, min_col=1, max_col=4):
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                            cell.border = thin_border
                    
                    for col in ws_teachers.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws_teachers.column_dimensions[column].width = adjusted_width

                # --- Нагрузка групп ---
                group_report_data = []
                for group_name in self.schedule[self.schedule['status'] == 'подтверждено']['group_name'].unique():
                    group_schedule = self.schedule[
                        (self.schedule['group_name'] == group_name) &
                        (self.schedule['status'] == 'подтверждено')
                    ]
                    teachers = ', '.join(sorted(group_schedule['teacher_name'].unique()))
                    hours = len(group_schedule)

                    # Формируем строку с предметами и часами
                    subject_hours = group_schedule['subject_name'].value_counts().to_dict()
                    fact_subjects_str = ""
                    for subj, count in subject_hours.items():
                        fact_subjects_str += f"{subj}: {count} ч\n"  # Используем \n для Excel
                    
                    if not fact_subjects_str:
                        fact_subjects_str = "—"

                    group_report_data.append([group_name, teachers, fact_subjects_str, hours])

                if group_report_data:
                    ws_groups = wb.create_sheet('Нагрузка групп')
                    ws_groups.append(['Группа', 'Преподаватели', 'Предметы', 'Часы'])
                    
                    for col_num in range(1, 5):
                        cell = ws_groups.cell(row=1, column=col_num)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    for row_data in group_report_data:
                        ws_groups.append(row_data)
                    
                    for row in ws_groups.iter_rows(min_row=2, max_row=ws_groups.max_row, min_col=1, max_col=4):
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                            cell.border = thin_border
                    
                    for col in ws_groups.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws_groups.column_dimensions[column].width = adjusted_width

            # Сохраняем файл
            wb.save(filename)
            QMessageBox.information(self, "Экспорт", f"Расписание успешно экспортировано в {filename}")
            self.create_backup()

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка экспорта: {str(e)}")
            
            
    def export_to_website(self):
        if self.schedule.empty:
            QMessageBox.information(self, "Информация", "Сначала сгенерируйте расписание")
            return
        filename, _ = QFileDialog.getSaveFileName(self, "Сохранить расписание как веб-сайт", "", "HTML files (*.html);;All files (*.*)")
        if not filename:
            return
        try:
            confirmed_schedule = self.schedule[self.schedule['status'] == 'подтверждено']
            if confirmed_schedule.empty:
                QMessageBox.warning(self, "Предупреждение", "Нет подтвержденных занятий для экспорта")
                return
            unique_groups = sorted(confirmed_schedule['group_name'].unique())
            unique_teachers = sorted(confirmed_schedule['teacher_name'].unique())
            unique_classrooms = sorted(confirmed_schedule['classroom_name'].unique())
            days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье'][:self.settings['days_per_week']]
            def parse_time(time_str):
                start_time_str = time_str.split('-')[0].strip()
                return dt_datetime.strptime(start_time_str, '%H:%M')
            html_content = f"""
    <!DOCTYPE html>
    <html lang="ru">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>📅 Расписание занятий - {self.settings.get('school_name', 'Образовательное учреждение')}</title>
        <style>
            * {{
                box-sizing: border-box;
                margin: 0;
                padding: 0;
            }}
            body {{
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                line-height: 1.6;
                color: #333;
                background-color: #f5f7fa;
                padding: 20px;
            }}
            .container {{
                max-width: 1200px;
                margin: 0 auto;
                background: white;
                border-radius: 10px;
                box-shadow: 0 0 20px rgba(0,0,0,0.1);
                overflow: hidden;
            }}
            .header {{
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                padding: 30px 20px;
                text-align: center;
            }}
            .header h1 {{
                margin: 0 0 10px;
                font-size: 2.5em;
            }}
            .header p {{
                margin: 0;
                font-size: 1.2em;
                opacity: 0.9;
            }}
            .filters {{
                padding: 20px;
                background-color: #2c3e50;
                color: white;
                display: flex;
                flex-wrap: wrap;
                gap: 15px;
                align-items: center;
                justify-content: center;
            }}
            .filters label {{
                font-weight: bold;
                min-width: 120px;
                text-align: right;
            }}
            .filters select, .filters button {{
                padding: 8px 12px;
                border: none;
                border-radius: 5px;
                font-size: 14px;
            }}
            .filters button {{
                background-color: #3498db;
                color: white;
                cursor: pointer;
                font-weight: bold;
                transition: background-color 0.3s;
            }}
            .filters button:hover {{
                background-color: #2980b9;
            }}
            .tabs {{
                display: flex;
                background-color: #ecf0f1;
                border-bottom: 2px solid #bdc3c7;
            }}
            .tab {{
                padding: 15px 25px;
                cursor: pointer;
                font-weight: bold;
                color: #7f8c8d;
                border-bottom: 3px solid transparent;
                transition: all 0.3s;
            }}
            .tab.active {{
                color: #2c3e50;
                border-bottom: 3px solid #3498db;
                background-color: white;
            }}
            .tab-content {{
                display: none;
                padding: 20px;
            }}
            .tab-content.active {{
                display: block;
            }}
            .schedule-table {{
                background: white;
                border-radius: 8px;
                box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                margin-bottom: 30px;
                overflow: hidden;
                transition: transform 0.2s;
            }}
            .schedule-table:hover {{
                transform: translateY(-2px);
                box-shadow: 0 4px 20px rgba(0,0,0,0.15);
            }}
            .schedule-title {{
                background-color: #34495e;
                color: white;
                padding: 15px 20px;
                font-weight: bold;
                font-size: 1.3em;
                text-align: center;
                border-bottom: 2px solid #2c3e50;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin: 0;
            }}
            th, td {{
                padding: 15px;
                text-align: center;
                border: 1px solid #e0e0e0;
                vertical-align: top;
            }}
            th {{
                background-color: #ecf0f1;
                color: #2c3e50;
                font-weight: bold;
                position: sticky;
                top: 0;
                z-index: 10;
                transition: background-color 0.3s;
            }}
            tr:nth-child(even) {{
                background-color: #f8f9fa;
            }}
            tr:hover {{
                background-color: #e3f2fd;
            }}
            .lesson {{
                font-size: 0.95em;
                font-weight: 500;
            }}
            .lesson-info {{
                font-size: 0.8em;
                color: #7f8c8d;
                margin-top: 5px;
                display: block;
            }}
            .no-lesson {{
                color: #95a5a6;
                font-style: italic;
                padding: 20px 0;
            }}
            .footer {{
                text-align: center;
                padding: 20px;
                background-color: #2c3e50;
                color: #bdc3c7;
                font-size: 0.9em;
                margin-top: 20px;
            }}
            .today {{
                background-color: #e3f2fd !important;
                font-weight: bold;
            }}
            .today-header {{
                background-color: #2196F3 !important;
                color: white !important;
            }}
            .tomorrow {{
                background-color: #f0f8ff !important;
            }}
            .tomorrow-header {{
                background-color: #03A9F4 !important;
                color: white !important;
            }}
            @media (max-width: 768px) {{
                body {{
                    padding: 10px;
                }}
                .header h1 {{
                    font-size: 2em;
                }}
                .filters {{
                    flex-direction: column;
                    align-items: stretch;
                }}
                .filters label {{
                    text-align: left;
                    margin-bottom: 5px;
                }}
                th, td {{
                    padding: 10px 5px;
                    font-size: 0.9em;
                }}
                .lesson-info {{
                    font-size: 0.75em;
                }}
                .tab {{
                    padding: 10px 15px;
                    font-size: 0.9em;
                }}
            }}
            @media (max-width: 480px) {{
                .header h1 {{
                    font-size: 1.5em;
                }}
                th, td {{
                    padding: 8px 4px;
                    font-size: 0.8em;
                }}
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>📅 Расписание занятий</h1>
                <p>{self.settings.get('school_name', 'Образовательное учреждение')}</p>
                <p>Учебный год: {self.settings.get('academic_year', 'Не указан')}</p>
            </div>
            <div class="filters">
                <label for="groupFilter">Группа:</label>
                <select id="groupFilter">
                    <option value="">Все группы</option>
                    {''.join(f'<option value="{group}">{group}</option>' for group in unique_groups)}
                </select>
                <label for="teacherFilter">Преподаватель:</label>
                <select id="teacherFilter">
                    <option value="">Все преподаватели</option>
                    {''.join(f'<option value="{teacher}">{teacher}</option>' for teacher in unique_teachers)}
                </select>
                <label for="classroomFilter">Аудитория:</label>
                <select id="classroomFilter">
                    <option value="">Все аудитории</option>
                    {''.join(f'<option value="{classroom}">{classroom}</option>' for classroom in unique_classrooms)}
                </select>
                <button onclick="printSchedule()">🖨️ Печать</button>
            </div>
            <div class="tabs">
                <div class="tab active" onclick="openTab(event, 'groupsTab')">По группам</div>
                <div class="tab" onclick="openTab(event, 'teachersTab')">По преподавателям</div>
            </div>
            <div id="groupsTab" class="tab-content active">
                <div class="schedule-container" id="scheduleContainerGroups">
    """
            for group_name in unique_groups:
                group_schedule = confirmed_schedule[confirmed_schedule['group_name'] == group_name]
                times = sorted(group_schedule['time'].unique(), key=parse_time)
                html_content += f"""
                    <div class="schedule-table" data-group="{group_name}" data-teacher="" data-classroom="">
                        <div class="schedule-title">Группа: {group_name}</div>
                        <table>
                            <thead>
                                <tr>
                                    <th>Время</th>
                                    {''.join(f'<th class="day-header" data-day="{day}">{day}</th>' for day in days)}
                                </tr>
                            </thead>
                            <tbody>
    """
                for time_slot in times:
                    html_content += f"                        <tr>\n                            <td><strong>{time_slot}</strong></td>\n"
                    for day in days:
                        lesson = group_schedule[
                            (group_schedule['time'] == time_slot) &
                            (group_schedule['day'] == day)
                        ]
                        if not lesson.empty:
                            lesson_info = lesson.iloc[0]
                            lesson_html = f"""
                                <div class="lesson" data-teacher="{lesson_info['teacher_name']}" data-classroom="{lesson_info['classroom_name']}">
                                    {lesson_info['subject_name']}
                                    <span class="lesson-info">
                                        {lesson_info['teacher_name']}<br>
                                        {lesson_info['classroom_name']}
                                    </span>
                                </div>
                            """.strip().replace('\n', '').replace('    ', '')
                        else:
                            lesson_html = '<span class="no-lesson">—</span>'
                        html_content += f"                            <td class='day-cell' data-day='{day}'>{lesson_html}</td>\n"
                    html_content += "                        </tr>\n"
                html_content += """                    </tbody>
                    </table>
                </div>
    """
            html_content += """
                </div>
            </div>
            <div id="teachersTab" class="tab-content">
                <div class="schedule-container" id="scheduleContainerTeachers">
    """
            for teacher_name in unique_teachers:
                teacher_schedule = confirmed_schedule[confirmed_schedule['teacher_name'] == teacher_name]
                times = sorted(teacher_schedule['time'].unique(), key=parse_time)
                html_content += f"""
                    <div class="schedule-table" data-group="" data-teacher="{teacher_name}" data-classroom="">
                        <div class="schedule-title">Преподаватель: {teacher_name}</div>
                        <table>
                            <thead>
                                <tr>
                                    <th>Время</th>
                                    {''.join(f'<th class="day-header" data-day="{day}">{day}</th>' for day in days)}
                                </tr>
                            </thead>
                            <tbody>
    """
                for time_slot in times:
                    html_content += f"                        <tr>\n                            <td><strong>{time_slot}</strong></td>\n"
                    for day in days:
                        lesson = teacher_schedule[
                            (teacher_schedule['time'] == time_slot) &
                            (teacher_schedule['day'] == day)
                        ]
                        if not lesson.empty:
                            lesson_info = lesson.iloc[0]
                            lesson_html = f"""
                                <div class="lesson" data-group="{lesson_info['group_name']}" data-classroom="{lesson_info['classroom_name']}">
                                    {lesson_info['subject_name']}
                                    <span class="lesson-info">
                                        {lesson_info['group_name']}<br>
                                        {lesson_info['classroom_name']}
                                    </span>
                                </div>
                            """.strip().replace('\n', '').replace('    ', '')
                        else:
                            lesson_html = '<span class="no-lesson">—</span>'
                        html_content += f"                            <td class='day-cell' data-day='{day}'>{lesson_html}</td>\n"
                    html_content += "                        </tr>\n"
                html_content += """                    </tbody>
                    </table>
                </div>
    """
            html_content += f"""
                </div>
            </div>
            <div class="footer">
                <p>Расписание сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}</p>
                <p>Директор: {self.settings.get('director', 'Не указан')}</p>
            </div>
        </div>
        <script>
            function filterSchedule() {{
                const groupValue = document.getElementById('groupFilter').value;
                const teacherValue = document.getElementById('teacherFilter').value;
                const classroomValue = document.getElementById('classroomFilter').value;
                const groupTables = document.querySelectorAll('#groupsTab .schedule-table');
                groupTables.forEach(table => {{
                    const groupMatch = !groupValue || table.dataset.group === groupValue;
                    let hasMatchingLesson = false;
                    const lessonElements = table.querySelectorAll('.lesson');
                    for (let lesson of lessonElements) {{
                        const lessonTeacher = lesson.dataset.teacher || '';
                        const lessonClassroom = lesson.dataset.classroom || '';
                        const teacherMatch = !teacherValue || lessonTeacher === teacherValue;
                        const classroomMatch = !classroomValue || lessonClassroom === classroomValue;
                        if (teacherMatch && classroomMatch) {{
                            hasMatchingLesson = true;
                            break;
                        }}
                    }}
                    if (groupMatch && hasMatchingLesson) {{
                        table.style.display = 'block';
                    }} else {{
                        table.style.display = 'none';
                    }}
                }});
                const teacherTables = document.querySelectorAll('#teachersTab .schedule-table');
                teacherTables.forEach(table => {{
                    const teacherMatch = !teacherValue || table.dataset.teacher === teacherValue;
                    let hasMatchingLesson = false;
                    const lessonElements = table.querySelectorAll('.lesson');
                    for (let lesson of lessonElements) {{
                        const lessonGroup = lesson.dataset.group || '';
                        const lessonClassroom = lesson.dataset.classroom || '';
                        const groupMatch = !groupValue || lessonGroup === groupValue;
                        const classroomMatch = !classroomValue || lessonClassroom === classroomValue;
                        if (groupMatch && classroomMatch) {{
                            hasMatchingLesson = true;
                            break;
                        }}
                    }}
                    if (teacherMatch && hasMatchingLesson) {{
                        table.style.display = 'block';
                    }} else {{
                        table.style.display = 'none';
                    }}
                }});
            }}
            function printSchedule() {{
                window.print();
            }}
            function highlightCurrentDay() {{
                const now = new Date();
                const currentDayIndex = now.getDay();
                const dayNames = ['Воскресенье', 'Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота'];
                const currentDayName = dayNames[currentDayIndex];
                const tomorrowDayName = dayNames[(currentDayIndex + 1) % 7];
                document.querySelectorAll('#groupsTab th.day-header[data-day="' + currentDayName + '"]').forEach(header => header.classList.add('today-header'));
                document.querySelectorAll('#groupsTab td.day-cell[data-day="' + currentDayName + '"]').forEach(cell => cell.classList.add('today'));
                document.querySelectorAll('#groupsTab th.day-header[data-day="' + tomorrowDayName + '"]').forEach(header => header.classList.add('tomorrow-header'));
                document.querySelectorAll('#groupsTab td.day-cell[data-day="' + tomorrowDayName + '"]').forEach(cell => cell.classList.add('tomorrow'));
                document.querySelectorAll('#teachersTab th.day-header[data-day="' + currentDayName + '"]').forEach(header => header.classList.add('today-header'));
                document.querySelectorAll('#teachersTab td.day-cell[data-day="' + currentDayName + '"]').forEach(cell => cell.classList.add('today'));
                document.querySelectorAll('#teachersTab th.day-header[data-day="' + tomorrowDayName + '"]').forEach(header => header.classList.add('tomorrow-header'));
                document.querySelectorAll('#teachersTab td.day-cell[data-day="' + tomorrowDayName + '"]').forEach(cell => cell.classList.add('tomorrow'));
            }}
            function openTab(evt, tabName) {{
                const tabContents = document.getElementsByClassName("tab-content");
                for (let i = 0; i < tabContents.length; i++) {{
                    tabContents[i].classList.remove("active");
                }}
                const tabs = document.getElementsByClassName("tab");
                for (let i = 0; i < tabs.length; i++) {{
                    tabs[i].classList.remove("active");
                }}
                document.getElementById(tabName).classList.add("active");
                evt.currentTarget.classList.add("active");
            }}
            document.getElementById('groupFilter').addEventListener('change', filterSchedule);
            document.getElementById('teacherFilter').addEventListener('change', filterSchedule);
            document.getElementById('classroomFilter').addEventListener('change', filterSchedule);
            document.addEventListener('DOMContentLoaded', highlightCurrentDay);
        </script>
    </body>
    </html>
    """
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(html_content)
            QMessageBox.information(self, "Экспорт", f"Веб-сайт с расписанием успешно создан!\nФайл: {filename}\nТеперь вы можете открыть этот файл в браузере или загрузить его на ваш сайт.")
            self.create_backup()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка экспорта в HTML: {str(e)}")

    def add_holiday(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавить праздник")
        dialog.setModal(True)
        dialog.resize(400, 200)
        form_layout = QFormLayout(dialog)
        date_entry = QDateEdit()
        date_entry.setCalendarPopup(True)
        date_entry.setDate(datetime.now().date())
        name_entry = QLineEdit()
        type_combo = QComboBox()
        type_combo.addItems(["Государственный", "Учебный", "Каникулы"])
        form_layout.addRow("Дата:", date_entry)
        form_layout.addRow("Название:", name_entry)
        form_layout.addRow("Тип:", type_combo)
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(lambda: self._save_holiday(date_entry.date().toString("yyyy-MM-dd"), name_entry.text(), type_combo.currentText(), dialog))
        button_box.rejected.connect(dialog.reject)
        form_layout.addRow(button_box)
        dialog.exec_()

    def _save_holiday(self, date_str, name, holiday_type, dialog):
        if not date_str or not name:
            QMessageBox.warning(self, "Предупреждение", "Заполните все поля")
            return
        try:
            datetime.strptime(date_str, '%Y-%m-%d')
            self.holidays.append({
                'date': date_str,
                'name': name,
                'type': holiday_type
            })
            self.load_holidays_data()
            self.create_backup()
            dialog.accept()
        except ValueError:
            QMessageBox.critical(self, "Ошибка", "Неверный формат даты. Используйте ГГГГ-ММ-ДД")

    def delete_holiday(self):
        selected_items = self.holidays_tree.selectedItems()
        if not selected_items:
            QMessageBox.information(self, "Информация", "Выберите праздник для удаления")
            return
        if QMessageBox.question(self, "Подтверждение", "Удалить выбранный праздник?") == QMessageBox.Yes:
            row = selected_items[0].row()
            holiday_date = self.holidays_tree.item(row, 0).text()
            holiday_name = self.holidays_tree.item(row, 1).text()
            self.holidays = [h for h in self.holidays if not (h['date'] == holiday_date and h['name'] == holiday_name)]
            self.load_holidays_data()
            self.create_backup()

    def save_current_schedule(self):
        if self.schedule.empty:
            QMessageBox.warning(self, "Предупреждение", "Нет расписания для сохранения. Сначала сгенерируйте его.")
            return
        school_name = self.settings.get('school_name', 'Расписание').replace(" ", "_")
        academic_year = self.settings.get('academic_year', 'Год_не_указан').replace("/", "_")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{school_name}_{academic_year}_{timestamp}.json"
        filepath = os.path.join(self.archive_dir, filename)
        try:
            schedule_dict = self.schedule.to_dict(orient='records') if not self.schedule.empty else []
            archive_data = {
                'settings': self.settings,
                'groups': self.groups,
                'teachers': self.teachers,
                'classrooms': self.classrooms,
                'subjects': self.subjects,
                'schedule': schedule_dict,
                'holidays': self.holidays,
                'substitutions': self.substitutions,
                'saved_at': datetime.now().isoformat()
            }
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(archive_data, f, ensure_ascii=False, indent=2)
            QMessageBox.information(self, "Успех", f"Расписание успешно сохранено в архив! Файл: {filename}")
            self.load_archive_list()
            self.create_backup()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка сохранения расписания в архив: {str(e)}")

    def load_archived_schedule(self):
        selected_items = self.archive_tree.selectedItems()
        if not selected_items:
            QMessageBox.information(self, "Информация", "Выберите расписание для загрузки")
            return
        row = selected_items[0].row()
        filename = self.archive_tree.item(row, 0).text()
        filepath = os.path.join(self.archive_dir, filename)
        if QMessageBox.question(self, "Подтверждение", f"Вы уверены, что хотите загрузить расписание из файла {filename}?\nТекущие данные будут заменены.") == QMessageBox.Yes:
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                self.settings = data.get('settings', self.settings)
                self.groups = data.get('groups', [])
                self.teachers = data.get('teachers', [])
                self.classrooms = data.get('classrooms', [])
                self.subjects = data.get('subjects', [])
                self.holidays = data.get('holidays', [])
                self.substitutions = data.get('substitutions', [])
                schedule_data = data.get('schedule', [])
                if schedule_data:
                    self.schedule = pd.DataFrame(schedule_data)
                else:
                    self.schedule = pd.DataFrame()
                self.load_groups_data()
                self.load_teachers_data()
                self.load_classrooms_data()
                self.load_subjects_data()
                self.load_holidays_data()
                self.filter_schedule()
                self.update_reports()
                QMessageBox.information(self, "Успех", f"Расписание успешно загружено из {filename}")
                self.create_backup()
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка загрузки расписания: {str(e)}")

    def delete_archived_schedule(self):
        selected_items = self.archive_tree.selectedItems()
        if not selected_items:
            QMessageBox.information(self, "Информация", "Выберите расписание для удаления")
            return
        row = selected_items[0].row()
        filename = self.archive_tree.item(row, 0).text()
        filepath = os.path.join(self.archive_dir, filename)
        if QMessageBox.question(self, "Подтверждение", f"Вы уверены, что хотите удалить файл {filename}?") == QMessageBox.Yes:
            try:
                os.remove(filepath)
                self.load_archive_list()
                QMessageBox.information(self, "Успех", f"Расписание {filename} успешно удалено из архива")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка удаления: {str(e)}")

    def export_archived_schedule(self):
        selected_items = self.archive_tree.selectedItems()
        if not selected_items:
            QMessageBox.information(self, "Информация", "Выберите расписание для экспорта")
            return
        
        row = selected_items[0].row()
        filename = self.archive_tree.item(row, 0).text()
        filepath = os.path.join(self.archive_dir, filename)
        
        save_path, _ = QFileDialog.getSaveFileName(
            self, f"Экспорт расписания {filename}", "", 
            "Excel files (*.xlsx);;All files (*.*)"
        )
        
        if not save_path:
            return
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            schedule_data = data.get('schedule', [])
            if not schedule_data:
                QMessageBox.warning(self, "Предупреждение", "В выбранном файле нет данных расписания")
                return
            
            # Создаем книгу Excel
            wb = openpyxl.Workbook()
            
            # === Лист "Расписание по дням" (красивый вывод) ===
            ws_pretty = wb.create_sheet('Расписание по дням', 0)  # Первый лист
            
            # Стили
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            day_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            day_header_font = Font(bold=True, size=11)
            lesson_font = Font(size=10)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Группируем расписание по неделям и дням
            schedule_by_week_day = {}
            for lesson in schedule_data:
                week = lesson.get('week', '')
                day = lesson.get('day', '')
                time = lesson.get('time', '')
                
                if week not in schedule_by_week_day:
                    schedule_by_week_day[week] = {}
                if day not in schedule_by_week_day[week]:
                    schedule_by_week_day[week][day] = []
                
                schedule_by_week_day[week][day].append({
                    'time': time,
                    'group': lesson.get('group_name', ''),
                    'subject': lesson.get('subject_name', ''),
                    'teacher': lesson.get('teacher_name', ''),
                    'classroom': lesson.get('classroom_name', ''),
                    'status': lesson.get('status', '')
                })
            
            # Заголовки для красивого вывода
            pretty_headers = ['Время', 'Группа', 'Предмет', 'Преподаватель', 'Аудитория', 'Статус']
            
            current_row = 1
            
            # Сортируем недели и дни
            weeks = sorted(schedule_by_week_day.keys(), key=lambda x: int(x) if str(x).isdigit() else 0)
            
            for week in weeks:
                days_order = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
                
                for day in days_order:
                    if day in schedule_by_week_day[week]:
                        lessons = schedule_by_week_day[week][day]
                        if not lessons:
                            continue
                        
                        # Заголовок дня
                        day_title = f"Неделя {week}, {day}"
                        ws_pretty.merge_cells(f'A{current_row}:F{current_row}')
                        day_cell = ws_pretty.cell(row=current_row, column=1, value=day_title)
                        day_cell.fill = day_header_fill
                        day_cell.font = day_header_font
                        day_cell.alignment = Alignment(horizontal='center')
                        current_row += 1
                        
                        # Заголовки столбцов
                        for col, header in enumerate(pretty_headers, 1):
                            cell = ws_pretty.cell(row=current_row, column=col, value=header)
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.border = border
                        
                        current_row += 1
                        
                        # Данные занятий
                        for lesson in sorted(lessons, key=lambda x: x['time']):
                            ws_pretty.cell(row=current_row, column=1, value=lesson['time']).font = lesson_font
                            ws_pretty.cell(row=current_row, column=2, value=lesson['group']).font = lesson_font
                            ws_pretty.cell(row=current_row, column=3, value=lesson['subject']).font = lesson_font
                            ws_pretty.cell(row=current_row, column=4, value=lesson['teacher']).font = lesson_font
                            ws_pretty.cell(row=current_row, column=5, value=lesson['classroom']).font = lesson_font
                            
                            status_cell = ws_pretty.cell(row=current_row, column=6, value=lesson['status'])
                            status_cell.font = lesson_font
                            # Цвет статуса
                            if lesson['status'] == 'подтвержден':
                                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif lesson['status'] == 'свободно':
                                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            
                            # Границы для всей строки
                            for col in range(1, 7):
                                ws_pretty.cell(row=current_row, column=col).border = border
                            
                            current_row += 1
                        
                        # Пустая строка между днями
                        current_row += 1
            
            # Настройка ширины столбцов
            column_widths = [15, 12, 25, 20, 12, 12]
            for i, width in enumerate(column_widths, 1):
                ws_pretty.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
            # === Оригинальный лист "Расписание" (как было) ===
            ws_schedule = wb.create_sheet('Расписание (полное)')
            
            # Заголовки
            headers = ['Неделя', 'День', 'Время', 'Группа', 'Предмет', 'Преподаватель', 'Аудитория', 'Статус']
            ws_schedule.append(headers)
            
            # Данные расписания
            for lesson in schedule_data:
                row_data = [
                    lesson.get('week', ''),
                    lesson.get('day', ''),
                    lesson.get('time', ''),
                    lesson.get('group_name', ''),
                    lesson.get('subject_name', ''),
                    lesson.get('teacher_name', ''),
                    lesson.get('classroom_name', ''),
                    lesson.get('status', '')
                ]
                ws_schedule.append(row_data)
            
            # === Остальные листы (без изменений) ===
            # Лист "Группы"
            ws_groups = wb.create_sheet('Группы')
            groups_data = data.get('groups', [])
            if groups_data:
                headers = ['ID', 'Название', 'Тип', 'Студентов', 'Курс', 'Специальность']
                ws_groups.append(headers)
                
                for group in groups_data:
                    students_value = group.get('students', '')
                    
                    # ИСПРАВЛЕНИЕ: Проверяем тип данных
                    if isinstance(students_value, (int, float)):
                        # Если это уже число, оставляем как есть
                        pass
                    elif isinstance(students_value, str):
                        numbers = []
                        parts = students_value.split(',')
                        for part in parts:
                            if ':' in part:
                                number_part = part.split(':')[-1].strip()
                                # Безопасное извлечение цифр
                                number_str = ''.join(filter(str.isdigit, number_part))
                                if number_str:
                                    try:
                                        numbers.append(int(number_str))
                                    except ValueError:
                                        continue
                        
                        if numbers:
                            students_value = numbers[0]
                        else:
                            students_value = 0
                    else:
                        # Если какой-то другой тип, устанавливаем 0
                        students_value = 0
                    
                    row_data = [
                        group.get('id', ''),
                        group.get('name', ''),
                        group.get('type', ''),
                        students_value,
                        group.get('course', ''),
                        group.get('specialty', '')
                    ]
                    ws_groups.append(row_data)
            
            # Лист "Преподаватели"
            ws_teachers = wb.create_sheet('Преподаватели')
            teachers_data = data.get('teachers', [])
            if teachers_data:
                headers = ['ID', 'ФИО', 'Макс. часов', 'Квалификация', 'Стаж', 'Контакты']
                ws_teachers.append(headers)
                
                for teacher in teachers_data:
                    max_hours = self._safe_extract_number(teacher.get('max_hours', ''))
                    experience = self._safe_extract_number(teacher.get('experience', ''))
                    
                    row_data = [
                        teacher.get('id', ''),
                        teacher.get('name', ''),
                        max_hours,
                        teacher.get('qualification', ''),
                        experience,
                        teacher.get('contacts', '')
                    ]
                    ws_teachers.append(row_data)
            
            # Лист "Аудитории"
            ws_classrooms = wb.create_sheet('Аудитории')
            classrooms_data = data.get('classrooms', [])
            if classrooms_data:
                headers = ['ID', 'Номер', 'Вместимость', 'Тип', 'Оборудование', 'Расположение']
                ws_classrooms.append(headers)
                
                for classroom in classrooms_data:
                    capacity = self._safe_extract_number(classroom.get('capacity', ''))
                    
                    row_data = [
                        classroom.get('id', ''),
                        classroom.get('name', ''),
                        capacity,
                        classroom.get('type', ''),
                        classroom.get('equipment', ''),
                        classroom.get('location', '')
                    ]
                    ws_classrooms.append(row_data)
            
            # Лист "Предметы"
            ws_subjects = wb.create_sheet('Предметы')
            subjects_data = data.get('subjects', [])
            if subjects_data:
                headers = ['ID', 'Название', 'Тип группы', 'Часов/неделю', 'Форма контроля', 'Кафедра']
                ws_subjects.append(headers)
                
                for subject in subjects_data:
                    hours_per_week = self._safe_extract_number(subject.get('hours_per_week', ''))
                    
                    row_data = [
                        subject.get('id', ''),
                        subject.get('name', ''),
                        subject.get('group_type', ''),
                        hours_per_week,
                        subject.get('assessment', ''),
                        subject.get('department', '')
                    ]
                    ws_subjects.append(row_data)
            
            # Лист "Праздники"
            ws_holidays = wb.create_sheet('Праздники')
            holidays_data = data.get('holidays', [])
            if holidays_data:
                headers = ['Дата', 'Название', 'Тип']
                ws_holidays.append(headers)
                
                for holiday in holidays_data:
                    row_data = [
                        holiday.get('date', ''),
                        holiday.get('name', ''),
                        holiday.get('type', '')
                    ]
                    ws_holidays.append(row_data)
            
            # Удаляем пустой лист по умолчанию
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Применяем стили к заголовкам всех листов (кроме красивого)
            for sheet_name in wb.sheetnames:
                if sheet_name != 'Расписание по дням':
                    ws = wb[sheet_name]
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
            
            # Сохраняем файл
            wb.save(save_path)
            QMessageBox.information(self, "Экспорт", f"Расписание успешно экспортировано в {save_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка экспорта: {str(e)}")

    def _safe_extract_number(self, value):
        """Безопасное извлечение числа из значения"""
        if isinstance(value, (int, float)):
            return value
        elif isinstance(value, str):
            # Пробуем найти числа в строке
            numbers = []
            parts = value.split(',')
            for part in parts:
                if ':' in part:
                    number_part = part.split(':')[-1].strip()
                    number_str = ''.join(filter(lambda x: x.isdigit() or x == '.', number_part))
                    if number_str:
                        try:
                            numbers.append(float(number_str))
                        except ValueError:
                            continue
            
            if numbers:
                return numbers[0]
        
        return 0

    def _extract_number_from_string(self, text):
        """Извлекает число из строки вида '(Математика): 18, Информатика): 4'"""
        if not isinstance(text, str):
            return text
        
        numbers = []
        parts = text.split(',')
        for part in parts:
            if ':' in part:
                number_part = part.split(':')[-1].strip()
                number_str = ''.join(filter(lambda x: x.isdigit() or x == '.', number_part))
                if number_str:
                    try:
                        numbers.append(float(number_str))
                    except ValueError:
                        continue
        
        if numbers:
            return numbers[0]
        else:
            return 0
            

    def export_group_schedule_to_excel(self):
        """Экспорт расписания по группам в отдельные листы Excel"""
        if self.schedule.empty:
            QMessageBox.information(self, "Информация", "Нет расписания для экспорта")
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self, "Экспорт расписания по группам", "", 
            "Excel files (*.xlsx);;All files (*.*)"
        )
        
        if not filename:
            return
        
        try:
            confirmed_schedule = self.schedule[self.schedule['status'] == 'подтверждено']
            if confirmed_schedule.empty:
                QMessageBox.warning(self, "Предупреждение", "Нет подтвержденных занятий для экспорта")
                return

            # Создаем книгу Excel
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Удаляем пустой лист по умолчанию

            # Стили для оформления
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            day_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            day_header_font = Font(bold=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))

            # Получаем уникальные группы
            unique_groups = sorted(confirmed_schedule['group_name'].unique())
            
            for group_name in unique_groups:
                # Создаем лист для каждой группы
                ws = wb.create_sheet(f"Группа {group_name}")
                
                # Получаем расписание для текущей группы
                group_schedule = confirmed_schedule[confirmed_schedule['group_name'] == group_name]
                
                # Сортируем по неделям и дням
                weeks = sorted(group_schedule['week'].unique())
                days_order = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
                
                current_row = 1
                
                for week in weeks:
                    # Заголовок недели
                    ws.merge_cells(f'A{current_row}:G{current_row}')
                    week_cell = ws.cell(row=current_row, column=1, value=f"Неделя {week} - Группа {group_name}")
                    week_cell.fill = day_header_fill
                    week_cell.font = day_header_font
                    week_cell.alignment = Alignment(horizontal='center')
                    current_row += 1
                    
                    for day in days_order:
                        day_schedule = group_schedule[
                            (group_schedule['week'] == week) & 
                            (group_schedule['day'] == day)
                        ]
                        
                        if not day_schedule.empty:
                            # Заголовок дня
                            ws.merge_cells(f'A{current_row}:G{current_row}')
                            day_cell = ws.cell(row=current_row, column=1, value=day)
                            day_cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                            day_cell.font = Font(bold=True)
                            day_cell.alignment = Alignment(horizontal='center')
                            current_row += 1
                            
                            # Заголовки столбцов
                            headers = ['Время', 'Предмет', 'Преподаватель', 'Аудитория', 'Тип занятия', 'Статус', 'Примечания']
                            for col, header in enumerate(headers, 1):
                                cell = ws.cell(row=current_row, column=col, value=header)
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.border = border
                            current_row += 1
                            
                            # Данные занятий (сортируем по времени)
                            day_lessons = day_schedule.sort_values('time')
                            for _, lesson in day_lessons.iterrows():
                                ws.cell(row=current_row, column=1, value=lesson['time']).border = border
                                ws.cell(row=current_row, column=2, value=lesson['subject_name']).border = border
                                ws.cell(row=current_row, column=3, value=lesson['teacher_name']).border = border
                                ws.cell(row=current_row, column=4, value=lesson['classroom_name']).border = border
                                ws.cell(row=current_row, column=5, value="Лекция/Практика").border = border
                                ws.cell(row=current_row, column=6, value=lesson['status']).border = border
                                ws.cell(row=current_row, column=7, value="").border = border
                                current_row += 1
                            
                            # Пустая строка между днями
                            current_row += 1

                # Настройка ширины столбцов
                column_widths = [15, 30, 25, 15, 15, 12, 20]
                for i, width in enumerate(column_widths, 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

            # Сохраняем файл
            wb.save(filename)
            QMessageBox.information(self, "Экспорт", 
                                   f"Расписание по группам успешно экспортировано!\n"
                                   f"Создано листов: {len(unique_groups)}\n"
                                   f"Файл: {filename}")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка экспорта: {str(e)}")


    def open_settings(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Настройки приложения")
        dialog.setModal(True)
        dialog.resize(550, 700)
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        basic_frame = QGroupBox("Основные настройки")
        basic_layout = QFormLayout(basic_frame)
        school_name_var = QLineEdit(self.settings.get('school_name', ''))
        director_var = QLineEdit(self.settings.get('director', ''))
        academic_year_var = QLineEdit(self.settings.get('academic_year', ''))
        start_date_var = QLineEdit(self.settings.get('start_date', datetime.now().date().isoformat()))
        basic_layout.addRow("Название учреждения:", school_name_var)
        basic_layout.addRow("Директор:", director_var)
        basic_layout.addRow("Учебный год:", academic_year_var)
        basic_layout.addRow("Дата начала года:", start_date_var)
        scroll_layout.addWidget(basic_frame)
        schedule_frame = QGroupBox("Параметры расписания")
        schedule_layout = QFormLayout(schedule_frame)
        days_per_week_var = QSpinBox()
        days_per_week_var.setRange(1, 7)
        days_per_week_var.setValue(self.settings.get('days_per_week', 5))
        lessons_per_day_var = QSpinBox()
        lessons_per_day_var.setRange(1, 12)
        lessons_per_day_var.setValue(self.settings.get('lessons_per_day', 6))
        weeks_var = QSpinBox()
        weeks_var.setRange(1, 52)
        weeks_var.setValue(self.settings.get('weeks', 2))
        schedule_layout.addRow("Дней в неделю:", days_per_week_var)
        schedule_layout.addRow("Занятий в день:", lessons_per_day_var)
        schedule_layout.addRow("Недель:", weeks_var)
        scroll_layout.addWidget(schedule_frame)
        bell_frame = QGroupBox("Расписание звонков")
        bell_layout = QFormLayout(bell_frame)
        bell_schedule_var = QLineEdit(self.settings.get('bell_schedule', '8:00-8:45,8:55-9:40,9:50-10:35,10:45-11:30,11:40-12:25,12:35-13:20'))
        bell_schedule_var.setReadOnly(True)
        bell_layout.addRow("Текущее расписание:", bell_schedule_var)
        open_editor_btn = QPushButton("⚙️ Открыть редактор...")
        bell_layout.addRow(open_editor_btn)
        bell_layout.addRow(QLabel("Редактор расписания звонков", font=QFont('Segoe UI', 9, QFont.StyleItalic)))
        scroll_layout.addWidget(bell_frame)
        backup_frame = QGroupBox("Настройки авто-бэкапа")
        backup_layout = QFormLayout(backup_frame)
        auto_backup_var = QCheckBox()
        auto_backup_var.setChecked(self.settings.get('auto_backup', True))
        backup_interval_var = QSpinBox()
        backup_interval_var.setRange(1, 1440)
        backup_interval_var.setValue(self.settings.get('backup_interval', 30))
        max_backups_var = QSpinBox()
        max_backups_var.setRange(1, 100)
        max_backups_var.setValue(self.settings.get('max_backups', 10))
        backup_layout.addRow("Авто-бэкап:", auto_backup_var)
        backup_layout.addRow("Интервал бэкапа (мин):", backup_interval_var)
        backup_layout.addRow("Макс. бэкапов:", max_backups_var)
        scroll_layout.addWidget(backup_frame)
        scroll_area.setWidget(scroll_content)
        main_layout = QVBoxLayout(dialog)
        main_layout.addWidget(scroll_area)
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(lambda: self._save_settings(
            school_name_var.text(), director_var.text(), academic_year_var.text(), start_date_var.text(),
            days_per_week_var.value(), lessons_per_day_var.value(), weeks_var.value(),
            auto_backup_var.isChecked(), backup_interval_var.value(), max_backups_var.value(),
            bell_schedule_var.text(), dialog))
        button_box.rejected.connect(dialog.reject)
        main_layout.addWidget(button_box)
        def open_editor_wrapper():
            current_schedule = bell_schedule_var.text()
            editor = BellScheduleEditor(dialog, current_schedule)
            if editor.exec_() == QDialog.Accepted and editor.result is not None:
                bell_schedule_var.setText(editor.result)
        open_editor_btn.clicked.connect(open_editor_wrapper)
        dialog.exec_()

    def _save_settings(self, school_name, director, academic_year, start_date,
                      days_per_week, lessons_per_day, weeks,
                      auto_backup, backup_interval, max_backups,
                      bell_schedule, dialog):
        self.settings['school_name'] = school_name
        self.settings['director'] = director
        self.settings['academic_year'] = academic_year
        self.settings['start_date'] = start_date
        self.settings['days_per_week'] = days_per_week
        self.settings['lessons_per_day'] = lessons_per_day
        self.settings['weeks'] = weeks
        self.settings['auto_backup'] = auto_backup
        self.settings['backup_interval'] = backup_interval
        self.settings['max_backups'] = max_backups
        self.settings['bell_schedule'] = bell_schedule
        self.restart_auto_backup()
        self.update_backup_indicator()
        dialog.accept()

    def open_backup_manager(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Менеджер бэкапов")
        dialog.setModal(True)
        dialog.resize(600, 400)
        layout = QVBoxLayout(dialog)
        label = QLabel("🛡️ Менеджер бэкапов")
        label.setFont(QFont('Arial', 14, QFont.Bold))
        label.setAlignment(Qt.AlignCenter)
        layout.addWidget(label)
        self.backup_tree = QTableWidget(0, 3)
        self.backup_tree.setHorizontalHeaderLabels(['Имя файла', 'Дата создания', 'Размер'])
        self.backup_tree.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.backup_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        layout.addWidget(self.backup_tree, 1)
        self.backup_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.backup_tree.customContextMenuRequested.connect(self.show_backup_context_menu)
        button_frame = QFrame()
        button_layout = QHBoxLayout(button_frame)
        refresh_btn = QPushButton("🔄 Обновить")
        refresh_btn.clicked.connect(self.load_backup_list)
        create_btn = QPushButton("💾 Создать бэкап")
        create_btn.clicked.connect(self.create_backup)
        restore_btn = QPushButton("📂 Восстановить")
        restore_btn.clicked.connect(self.restore_backup)
        delete_btn = QPushButton("🗑️ Удалить")
        delete_btn.clicked.connect(self.delete_backup)
        button_layout.addWidget(refresh_btn)
        button_layout.addWidget(create_btn)
        button_layout.addWidget(restore_btn)
        button_layout.addWidget(delete_btn)
        button_layout.addStretch()
        layout.addWidget(button_frame)
        self.load_backup_list()
        dialog.exec_()

    def show_backup_context_menu(self, position):
        menu = QMenu()
        restore_action = QAction("📂 Восстановить", self)
        delete_action = QAction("🗑️ Удалить", self)
        refresh_action = QAction("🔄 Обновить список", self)
        menu.addAction(restore_action)
        menu.addAction(delete_action)
        menu.addSeparator()
        menu.addAction(refresh_action)
        action = menu.exec_(self.backup_tree.mapToGlobal(position))
        if action == restore_action:
            self.restore_backup()
        elif action == delete_action:
            self.delete_backup()
        elif action == refresh_action:
            self.load_backup_list()

    def load_backup_list(self):
        self.backup_tree.setRowCount(0)
        try:
            backup_files = [f for f in os.listdir(self.backup_dir) if f.endswith('.zip')]
            backup_files.sort(reverse=True)
            for filename in backup_files:
                filepath = os.path.join(self.backup_dir, filename)
                stat = os.stat(filepath)
                creation_time = datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S')
                size = stat.st_size
                if size < 1024:
                    size_str = f"{size} B"
                elif size < 1024**2:
                    size_str = f"{size/1024:.1f} KB"
                else:
                    size_str = f"{size/(1024**2):.1f} MB"
                row = self.backup_tree.rowCount()
                self.backup_tree.insertRow(row)
                self.backup_tree.setItem(row, 0, QTableWidgetItem(filename))
                self.backup_tree.setItem(row, 1, QTableWidgetItem(creation_time))
                self.backup_tree.setItem(row, 2, QTableWidgetItem(size_str))
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка загрузки списка бэкапов: {e}")

    def restore_backup(self):
        selected_items = self.backup_tree.selectedItems()
        if not selected_items:
            QMessageBox.information(self, "Информация", "Выберите бэкап для восстановления")
            return
        row = selected_items[0].row()
        filename = self.backup_tree.item(row, 0).text()
        filepath = os.path.join(self.backup_dir, filename)
        if QMessageBox.question(self, "Подтверждение", f"Вы уверены, что хотите восстановить данные из {filename}?\nТекущие данные будут потеряны.") == QMessageBox.Yes:
            try:
                with zipfile.ZipFile(filepath, 'r') as zip_ref:
                    zip_ref.extractall('.')
                QMessageBox.information(self, "Успех", f"Данные успешно восстановлены из {filename}")
                self.load_data()
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка восстановления: {e}")

    def delete_backup(self):
        selected_items = self.backup_tree.selectedItems()
        if not selected_items:
            QMessageBox.information(self, "Информация", "Выберите бэкап для удаления")
            return
        row = selected_items[0].row()
        filename = self.backup_tree.item(row, 0).text()
        filepath = os.path.join(self.backup_dir, filename)
        if QMessageBox.question(self, "Подтверждение", f"Вы уверены, что хотите удалить {filename}?") == QMessageBox.Yes:
            try:
                os.remove(filepath)
                self.load_backup_list()
                QMessageBox.information(self, "Успех", f"Бэкап {filename} успешно удален")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка удаления: {e}")

    def create_backup(self):
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"backup_{timestamp}.zip"
            backup_filepath = os.path.join(self.backup_dir, backup_filename)
            temp_data_file = "temp_data.json"
            data = {
                'settings': self.settings,
                'groups': self.groups,
                'teachers': self.teachers,
                'classrooms': self.classrooms,
                'subjects': self.subjects,
                'schedule': self.schedule.to_dict() if not self.schedule.empty else {},
                'substitutions': self.substitutions,
                'holidays': self.holidays
            }
            with open(temp_data_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            with zipfile.ZipFile(backup_filepath, 'w') as zipf:
                zipf.write(temp_data_file)
            os.remove(temp_data_file)
            self.cleanup_old_backups()
            self.last_backup_time = datetime.now()
            self.update_backup_indicator()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка создания бэкапа: {str(e)}")

    def cleanup_old_backups(self):
        try:
            backup_files = [f for f in os.listdir(self.backup_dir) if f.endswith('.zip')]
            backup_files.sort(key=lambda x: os.path.getctime(os.path.join(self.backup_dir, x)))
            while len(backup_files) > self.settings.get('max_backups', 10):
                oldest_file = backup_files.pop(0)
                os.remove(os.path.join(self.backup_dir, oldest_file))
        except Exception as e:
            pass

    def start_auto_backup(self):
        if self.settings.get('auto_backup', True):
            interval = self.settings.get('backup_interval', 30) * 60 * 1000
            self.backup_timer = QTimer()
            self.backup_timer.timeout.connect(self.auto_backup)
            self.backup_timer.start(interval)

    def auto_backup(self):
        self.create_backup()

    def restart_auto_backup(self):
        if self.backup_timer:
            self.backup_timer.stop()
        self.start_auto_backup()

    def update_backup_indicator(self):
        if self.settings.get('auto_backup', True):
            self.backup_status_label.setText("Авто-бэкап: ВКЛ")
            self.backup_status_label.setStyleSheet(f"color: {COLORS['success']};")
            if self.last_backup_time:
                next_backup = self.last_backup_time + timedelta(minutes=self.settings.get('backup_interval', 30))
                self.next_backup_time = next_backup
                self.backup_info_label.setText(f"Следующий: {next_backup.strftime('%H:%M')}")
            else:
                self.backup_info_label.setText("Следующий: --:--")
        else:
            self.backup_status_label.setText("Авто-бэкап: ВЫКЛ")
            self.backup_status_label.setStyleSheet(f"color: {COLORS['danger']};")
            self.backup_info_label.setText("")

    def reschedule_lesson(self):
        selected_indexes = self.schedule_view.selectionModel().selectedIndexes()
        if not selected_indexes:
            QMessageBox.information(self, "Информация", "Выберите занятие в таблице для переноса")
            return
        selected_index = selected_indexes[0]
        row = selected_index.row()
        col = selected_index.column()
        if col == 0:
            QMessageBox.warning(self, "Предупреждение", "Выберите ячейку с занятием (не время)")
            return
        cell_data = self.schedule_proxy_model.data(selected_index)
        if not cell_data or not cell_data.strip():
            QMessageBox.warning(self, "Предупреждение", "Выбранная ячейка пуста. Нет занятия для переноса.")
            return
        time_index = self.schedule_proxy_model.index(row, 0)
        time_slot = self.schedule_proxy_model.data(time_index)
        if not time_slot:
            QMessageBox.warning(self, "Ошибка", "Не удалось определить время занятия")
            return
        day_header = self.schedule_proxy_model.headerData(col, Qt.Horizontal, Qt.DisplayRole)
        if not day_header:
            QMessageBox.warning(self, "Ошибка", "Не удалось определить день недели")
            return
        selected_day = day_header.strip()
        week_text = self.week_var.currentText()
        try:
            selected_week = int(week_text.split()[1]) if "Неделя" in week_text else 1
        except (ValueError, IndexError):
            selected_week = 1
        target_lesson = self.schedule[
            (self.schedule['week'] == selected_week) &
            (self.schedule['day'] == selected_day) &
            (self.schedule['time'] == time_slot) &
            (self.schedule['status'] == 'подтверждено')
        ]
        if target_lesson.empty:
            QMessageBox.information(self, "Информация", "Выбранное занятие не найдено в расписании")
            return
        lesson_info = target_lesson.iloc[0]
        lesson_idx = target_lesson.index[0]
        dialog = QDialog(self)
        dialog.setWindowTitle("Перенос занятия")
        dialog.setModal(True)
        dialog.resize(400, 250)
        layout = QVBoxLayout(dialog)
        info_label = QLabel(f"""
        <b>Перенос занятия:</b>
        <br>Предмет: {lesson_info['subject_name']}
        <br>Группа: {lesson_info['group_name']}
        <br>Текущий день: {selected_day}
        <br>Текущее время: {time_slot}
        <br>Преподаватель: {lesson_info['teacher_name']}
        """)
        info_label.setWordWrap(True)
        info_label.setAlignment(Qt.AlignLeft)
        layout.addWidget(info_label)
        week_layout = QHBoxLayout()
        week_layout.addWidget(QLabel("Новая неделя:"))
        new_week_var = QComboBox()
        new_week_var.addItems([str(i) for i in range(1, self.settings['weeks'] + 1)])
        new_week_var.setCurrentText(str(selected_week))
        week_layout.addWidget(new_week_var)
        layout.addLayout(week_layout)
        day_layout = QHBoxLayout()
        day_layout.addWidget(QLabel("Новый день:"))
        new_day_var = QComboBox()
        new_day_var.addItems(['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье'][:self.settings['days_per_week']])
        new_day_var.setCurrentText(selected_day)
        day_layout.addWidget(new_day_var)
        layout.addLayout(day_layout)
        time_layout = QHBoxLayout()
        time_layout.addWidget(QLabel("Новое время:"))
        new_time_var = QComboBox()
        bell_schedule_str = self.settings.get('bell_schedule', '8:00-8:45,8:55-9:40,9:50-10:35,10:45-11:30,11:40-12:25,12:35-13:20')
        times = [slot.strip() for slot in bell_schedule_str.split(',')]
        new_time_var.addItems(times)
        new_time_var.setCurrentText(time_slot)
        time_layout.addWidget(new_time_var)
        layout.addLayout(time_layout)
        reason_label = QLabel("Причина переноса:")
        layout.addWidget(reason_label)
        reason_entry = QTextEdit()
        reason_entry.setPlaceholderText("Укажите причину переноса (ремонт, мероприятие и т.д.)")
        reason_entry.setMaximumHeight(60)
        layout.addWidget(reason_entry)
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)
        if dialog.exec_() == QDialog.Accepted:
            new_week = int(new_week_var.currentText())
            new_day = new_day_var.currentText()
            new_time = new_time_var.currentText()
            reason_text = reason_entry.toPlainText().strip()
            if not reason_text:
                QMessageBox.warning(self, "Предупреждение", "Пожалуйста, укажите причину переноса.")
                return
            existing_lesson = self.schedule[
                (self.schedule['week'] == new_week) &
                (self.schedule['day'] == new_day) &
                (self.schedule['time'] == new_time) &
                (self.schedule['group_id'] == lesson_info['group_id']) &
                (self.schedule['status'] != 'свободно') &
                (self.schedule.index != lesson_idx)
            ]
            if not existing_lesson.empty:
                if QMessageBox.question(self, "Подтверждение", "В выбранное время у этой группы уже есть занятие. Заменить его?") != QMessageBox.Yes:
                    return
            target_row = self.schedule[
                (self.schedule['week'] == new_week) &
                (self.schedule['day'] == new_day) &
                (self.schedule['time'] == new_time) &
                (self.schedule['group_id'] == lesson_info['group_id'])
            ]
            if not target_row.empty:
                update_idx = target_row.index[0]
            else:
                QMessageBox.critical(self, "Ошибка", "Не удалось найти подходящий слот в расписании для переноса")
                return
            self.schedule.loc[update_idx, 'week'] = new_week
            self.schedule.loc[update_idx, 'day'] = new_day
            self.schedule.loc[update_idx, 'time'] = new_time
            self.schedule.loc[update_idx, 'subject_id'] = lesson_info['subject_id']
            self.schedule.loc[update_idx, 'subject_name'] = lesson_info['subject_name']
            self.schedule.loc[update_idx, 'teacher_id'] = lesson_info['teacher_id']
            self.schedule.loc[update_idx, 'teacher_name'] = lesson_info['teacher_name']
            self.schedule.loc[update_idx, 'classroom_id'] = lesson_info['classroom_id']
            self.schedule.loc[update_idx, 'classroom_name'] = lesson_info['classroom_name']
            self.schedule.loc[update_idx, 'status'] = 'подтверждено'
            self.schedule.loc[lesson_idx, ['subject_id', 'subject_name', 'teacher_id', 'teacher_name', 'classroom_id', 'classroom_name']] = [None, '', None, '', None, '']
            self.schedule.loc[lesson_idx, 'status'] = 'свободно'
            self.substitutions.append({
                'date': datetime.now().isoformat(),
                'week': selected_week,
                'day': selected_day,
                'time': time_slot,
                'new_week': new_week,
                'new_day': new_day,
                'new_time': new_time,
                'group': lesson_info['group_name'],
                'subject': lesson_info['subject_name'],
                'teacher': lesson_info['teacher_name'],
                'action': 'перенос',
                'reason': reason_text
            })
            self.filter_schedule()
            self.update_reports()
            self.create_backup()
            QMessageBox.information(self, "Успех", f"Занятие успешно перенесено!\nНовое время: {new_day} {new_time} (Неделя {new_week})\nПричина: {reason_text}")

    def check_and_update_experience(self):
        current_year = datetime.now().year
        last_update_year = self.settings.get('last_academic_year_update', current_year)
        if current_year > last_update_year:
            for teacher in self.teachers:
                teacher['experience'] = teacher.get('experience', 0) + (current_year - last_update_year)
            self.settings['last_academic_year_update'] = current_year
            self.load_teachers_data()
            self.create_backup()

    def find_free_slot(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Найти свободное время")
        dialog.setModal(True)
        dialog.resize(400, 250)
        layout = QVBoxLayout(dialog)
        search_type_var = QComboBox()
        search_type_var.addItems(["Группа", "Преподаватель", "Аудитория"])
        element_var = QComboBox()
        layout.addWidget(QLabel("Выберите тип:"))
        layout.addWidget(search_type_var)
        layout.addWidget(QLabel("Выберите элемент:"))
        layout.addWidget(element_var)
        def update_combo():
            search_type = search_type_var.currentText()
            if search_type == "Группа":
                values = [g['name'] for g in self.groups]
            elif search_type == "Преподаватель":
                values = [t['name'] for t in self.teachers]
            else:
                values = [c['name'] for c in self.classrooms]
            element_var.clear()
            element_var.addItems(values)
        search_type_var.currentIndexChanged.connect(update_combo)
        update_combo()
        def search_slot():
            search_type = search_type_var.currentText()
            element_name = element_var.currentText()
            if not element_name:
                QMessageBox.warning(self, "Предупреждение", "Выберите элемент")
                return
            free_slots = self.schedule[self.schedule['status'] == 'свободно']
            if search_type == "Группа":
                free_slots = free_slots[free_slots['group_name'] == element_name]
            elif search_type == "Преподаватель":
                free_slots = free_slots[free_slots['teacher_name'] == '']
            else:
                free_slots = free_slots[free_slots['classroom_name'] == '']
            if not free_slots.empty:
                first_slot = free_slots.iloc[0]
                QMessageBox.information(self, "Свободный слот", f"Ближайший свободный слот:\nНеделя: {first_slot['week']}\nДень: {first_slot['day']}\nВремя: {first_slot['time']}\nГруппа: {first_slot['group_name']}")
            else:
                QMessageBox.information(self, "Информация", "Свободных слотов не найдено")
        search_btn = QPushButton("Найти")
        search_btn.clicked.connect(search_slot)
        layout.addWidget(search_btn)
        dialog.exec_()

    def show_about(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("О программе")
        dialog.setModal(True)
        dialog.setFixedSize(605, 650)
        dialog.setWindowFlags(dialog.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        logo_label = QLabel("🎓")
        logo_label.setFont(QFont('Arial', 48, QFont.Bold))
        logo_label.setAlignment(Qt.AlignCenter)
        scroll_layout.addWidget(logo_label)
        title_label = QLabel("Система автоматического составления расписания")
        title_label.setFont(QFont('Arial', 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        scroll_layout.addWidget(title_label)
        version_label = QLabel("Версия 2.0")
        version_label.setFont(QFont('Arial', 11, QFont.Bold))
        version_label.setAlignment(Qt.AlignCenter)
        scroll_layout.addWidget(version_label)
        description_text = (
            "Это профессиональное приложение предназначено для автоматизации и управления\n"
            "процессом составления учебного расписания в образовательных учреждениях.\n"
            "Программа объединяет в себе мощный функционал автоматической генерации и гибкие\n"
            "инструменты ручного управления, что делает ее незаменимым помощником для\n"
            "администраторов и завучей."
        )
        desc_label = QLabel(description_text)
        desc_label.setWordWrap(True)
        desc_label.setAlignment(Qt.AlignLeft)
        scroll_layout.addWidget(desc_label)
        features_title = QLabel("🔑 Основные возможности:")
        features_title.setFont(QFont('Arial', 12, QFont.Bold))
        features_title.setAlignment(Qt.AlignLeft)
        scroll_layout.addWidget(features_title)
        features_text = (
            " 🚀 Автоматическая генерация расписания на основе заданных параметров\n"
            " ✍️ Полный ручной контроль: добавление, редактирование и удаление занятий\n"
            " 🔄 Управление заменами преподавателей с ведением журнала\n"
            " 📅 Визуальный календарь для удобного просмотра расписания\n"
            " 📊 Комплексная система отчетов по нагрузке преподавателей и групп\n"
            " 🧩 Проверка и автоматическая оптимизация для устранения конфликтов\n"
            " 💾 Архивирование и загрузка ранее сгенерированных расписаний\n"
            " 🌐 Экспорт расписания в HTML для публикации на сайте учреждения\n"
            " 📤 Экспорт данных и отчетов в формат Excel\n"
            " 🛡️ Автоматическое и ручное создание резервных копий данных\n"
            " 🎓 Управление базами данных: группы, преподаватели, аудитории, предметы, праздники"
        )
        features_label = QLabel(features_text)
        features_label.setWordWrap(True)
        features_label.setAlignment(Qt.AlignLeft)
        scroll_layout.addWidget(features_label)
        dev_frame = QFrame()
        dev_layout = QVBoxLayout(dev_frame)
        dev_layout.addWidget(QLabel("👨‍💻 Разработчик:", font=QFont('Arial', 10, QFont.Bold)))
        dev_layout.addWidget(QLabel("Команда разработчиков Mikhail Lukomskiy"))
        dev_layout.addWidget(QLabel("📧 Контактная почта: support@lukomsky.ru"))
        dev_layout.addWidget(QLabel("🌐 Официальный сайт: www.lukomsky.ru"))
        dev_layout.addWidget(QLabel("📅 Год выпуска: 2025"))
        scroll_layout.addWidget(dev_frame)
        school_frame = QFrame()
        school_layout = QVBoxLayout(school_frame)
        school_layout.addWidget(QLabel("🏫 Текущее учреждение:", font=QFont('Arial', 10, QFont.Bold)))
        school_layout.addWidget(QLabel(f"{self.settings.get('school_name', 'Не указано')}"))
        school_layout.addWidget(QLabel(f"Директор: {self.settings.get('director', 'Не указан')}"))
        school_layout.addWidget(QLabel(f"Учебный год: {self.settings.get('academic_year', 'Не указан')}"))
        scroll_layout.addWidget(school_frame)
        scroll_area.setWidget(scroll_content)
        main_layout = QVBoxLayout(dialog)
        main_layout.addWidget(scroll_area)
        close_button = QPushButton("Закрыть")
        close_button.clicked.connect(dialog.accept)
        main_layout.addWidget(close_button)
        dialog.exec_()

    def save_data(self):
        data = {
            'settings': self.settings,
            'groups': self.groups,
            'teachers': self.teachers,
            'classrooms': self.classrooms,
            'subjects': self.subjects,
            'substitutions': self.substitutions,
            'holidays': self.holidays
        }
        filename, _ = QFileDialog.getSaveFileName(self, "Сохранить данные как", "", "JSON files (*.json);;All files (*.*)")
        if filename:
            try:
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                QMessageBox.information(self, "Сохранение", "Данные успешно сохранены!")
                self.create_backup()
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка сохранения: {str(e)}")

    def load_data(self):
        filename, _ = QFileDialog.getOpenFileName(self, "Загрузить данные", "", "JSON files (*.json);;All files (*.*)")
        if not filename:
            return
        try:
            with open(filename, 'r', encoding='utf-8') as f:
                data = json.load(f)
            self.settings = data.get('settings', self.settings)
            self.groups = data.get('groups', [])
            self.teachers = data.get('teachers', [])
            self.classrooms = data.get('classrooms', [])
            self.subjects = data.get('subjects', [])
            self.holidays = data.get('holidays', [])
            self.substitutions = data.get('substitutions', [])
            schedule_data = data.get('schedule', [])
            if isinstance(schedule_data, list):
                expected_columns = ['week', 'day', 'time', 'group_id', 'group_name', 'subject_id', 'subject_name',
                                    'teacher_id', 'teacher_name', 'classroom_id', 'classroom_name', 'status']
                self.schedule = pd.DataFrame(columns=expected_columns)
                for item in schedule_data:
                    row = {}
                    for col in expected_columns:
                        row[col] = item.get(col, None)
                    self.schedule = self.schedule._append(row, ignore_index=True)
            else:
                self.schedule = pd.DataFrame()
            self.load_groups_data()
            self.load_teachers_data()
            self.load_classrooms_data()
            self.load_subjects_data()
            self.load_holidays_data()
            self.filter_schedule()
            self.update_reports()
            QMessageBox.information(self, "Успех", f"Данные успешно загружены из {filename}")
            self.create_backup()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка загрузки данных: {str(e)}")

    def open_substitutions(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Журнал замен и переносов")
        dialog.setModal(True)
        dialog.resize(1100, 400)
        layout = QVBoxLayout(dialog)
        self.substitutions_tree = QTableWidget(0, 11)
        self.substitutions_tree.setHorizontalHeaderLabels([
            'Дата', 'Действие', 'Группа', 'Предмет',
            'Старая дата', 'Старое время',
            'Новая дата', 'Новое время',
            'Преподаватель', 'Причина', 'Подробнее'
        ])
        self.substitutions_tree.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.substitutions_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        layout.addWidget(self.substitutions_tree, 1)
        self.substitutions_tree.setRowCount(0)
        for sub in self.substitutions:
            row = self.substitutions_tree.rowCount()
            self.substitutions_tree.insertRow(row)
            try:
                dt = datetime.fromisoformat(sub.get('date', ''))
                display_date = dt.strftime('%d.%m.%Y %H:%M')
            except:
                display_date = sub.get('date', '')
            self.substitutions_tree.setItem(row, 0, QTableWidgetItem(display_date))
            action = sub.get('action', 'замена')
            self.substitutions_tree.setItem(row, 1, QTableWidgetItem(action))
            self.substitutions_tree.setItem(row, 2, QTableWidgetItem(sub.get('group', '')))
            self.substitutions_tree.setItem(row, 3, QTableWidgetItem(sub.get('subject', '')))
            old_week = sub.get('week', '')
            old_day = sub.get('day', '')
            old_time = sub.get('time', '')
            old_date_str = f"Неделя {old_week}, {old_day}" if old_week and old_day else ""
            self.substitutions_tree.setItem(row, 4, QTableWidgetItem(old_date_str))
            self.substitutions_tree.setItem(row, 5, QTableWidgetItem(old_time))
            new_week = sub.get('new_week', '')
            new_day = sub.get('new_day', '')
            new_time = sub.get('new_time', '')
            new_date_str = f"Неделя {new_week}, {new_day}" if new_week and new_day else ""
            self.substitutions_tree.setItem(row, 6, QTableWidgetItem(new_date_str))
            self.substitutions_tree.setItem(row, 7, QTableWidgetItem(new_time))
            if action == 'замена':
                teacher_info = f"{sub.get('old_teacher', '')} → {sub.get('new_teacher', '')}"
            else:
                teacher_info = sub.get('teacher', '')
            self.substitutions_tree.setItem(row, 8, QTableWidgetItem(teacher_info))
            self.substitutions_tree.setItem(row, 9, QTableWidgetItem(sub.get('reason', 'Не указана')))
            self.substitutions_tree.setItem(row, 10, QTableWidgetItem(""))
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        dialog.exec_()

    def check_for_updates(self):
        import requests
        import webbrowser
        from packaging import version
        UPDATE_INFO_URL = "https://scheduleapp.lukomsky.ru/updates/latest_version.json"
        CURRENT_VERSION = "2.0"
        try:
            response = requests.get(UPDATE_INFO_URL, timeout=10)
            response.raise_for_status()
            update_info = response.json()
            latest_version = update_info.get('version', '0.0')
            download_url = update_info.get('download_url', '')
            changelog = update_info.get('changelog', 'Информация об изменениях отсутствует.')
            if version.parse(latest_version) > version.parse(CURRENT_VERSION):
                dialog = QDialog(self)
                dialog.setWindowTitle("Доступно обновление!")
                dialog.setModal(True)
                dialog.resize(500, 300)
                layout = QVBoxLayout(dialog)
                info_label = QLabel(f"""
                <h2>Доступна новая версия!</h2>
                <p><b>Текущая версия:</b> {CURRENT_VERSION}</p>
                <p><b>Новая версия:</b> {latest_version}</p>
                <p><b>Что нового:</b></p>
                <p>{changelog}</p>
                """)
                info_label.setWordWrap(True)
                layout.addWidget(info_label)
                button_box = QDialogButtonBox()
                download_btn = QPushButton("⬇️ Скачать обновление")
                skip_btn = QPushButton("Пропустить")
                button_box.addButton(download_btn, QDialogButtonBox.AcceptRole)
                button_box.addButton(skip_btn, QDialogButtonBox.RejectRole)
                download_btn.clicked.connect(lambda: webbrowser.open(download_url))
                skip_btn.clicked.connect(dialog.reject)
                layout.addWidget(button_box)
                dialog.exec_()
            else:
                QMessageBox.information(self, "Обновления", "У вас установлена последняя версия приложения.")
        except requests.exceptions.RequestException as req_err:
            QMessageBox.warning(self, "Ошибка сети", f"Не удалось проверить обновления: {str(req_err)}")
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Произошла ошибка при проверке обновлений: {str(e)}")

    def run(self):
        self.show()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ScheduleApp()
    window.run()
    sys.exit(app.exec_())
